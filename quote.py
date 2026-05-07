"""
MSM 견적 자동화
Usage: python3 quote.py

1. Open quote_input.xlsx
2. Type specs in column A (any format)
3. Save and close
4. Run: python3 quote.py
5. Open quote_input.xlsx again — prices filled in
"""
import os
import sys
import re
import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- Paths ---
DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(DIR, "data", "msm.sqlite")
EXCEL_PATH = os.path.join(DIR, "quote_input.xlsx")

# --- Parser: free text → structured spec ---

PRODUCT_ALIASES = [
    (r"Y[-\s]?STR(?:AINER)?", "Y-STRAINER"),
    (r"SW[-\s]?CHECK|SWING\s*CHECK", "SW-CHECK"),
    (r"CHECK\s*(?:V/?V|VALVE)?", "SW-CHECK"),
    (r"GLOBE\s*(?:V/?V|VALVE)?", "GLOBE"),
    (r"GATE\s*(?:V/?V|VALVE)?", "GATE"),
    (r"스트레이너|와이스트레이너", "Y-STRAINER"),
    (r"체크\s*밸브|스윙\s*체크", "SW-CHECK"),
    (r"글로브\s*밸브?", "GLOBE"),
    (r"게이트\s*밸브?", "GATE"),
]

RATING_MAP = {
    "10K": "10K", "20K": "20K",
    "150#": "10K", "150LB": "10K", "150": "10K",
    "300#": "20K", "300LB": "20K", "300": "20K",
}
HIGH_RATINGS = {"30K", "600#", "600LB", "900#", "900LB", "1500#", "1500LB"}

SIZE_INCH_TO_MM = {
    2: 50, 2.5: 65, 3: 80, 4: 100, 5: 125, 6: 150,
    8: 200, 10: 250, 12: 300, 14: 350, 16: 400, 18: 450, 20: 500,
}

SPECIAL_END_CONN = {"SW", "BW", "RTJ"}
SPECIAL_TRIM = {"HF", "STL"}


def parse_spec(raw: str) -> dict:
    """Parse free-form spec into structured fields."""
    q = raw.strip().upper()
    result = {
        "raw": raw.strip(),
        "body_type": None,
        "rating": None,
        "size_mm": None,
        "end_connection": None,
        "trim": None,
        "operation": None,
        "rfq_reasons": [],
    }

    # 1. Body type
    for pattern, value in PRODUCT_ALIASES:
        if re.search(pattern, q) or re.search(pattern, raw.strip()):
            result["body_type"] = value
            break

    # 2. Rating
    # Check high ratings first
    for hr in HIGH_RATINGS:
        if hr.replace("#", r"#?").replace("LB", r"(?:LB|#)") and hr in q:
            result["rating"] = hr
            result["rfq_reasons"].append(f"비표준 압력등급: {hr}")
            break
    if not result["rating"]:
        m = re.search(r"\b(\d{2,4})\s*(?:#|K|LB)", q)
        if m:
            key = m.group(0).strip().replace(" ", "")
            for alias, val in RATING_MAP.items():
                if alias in key or key.startswith(alias):
                    result["rating"] = val
                    break
            if not result["rating"] and key in HIGH_RATINGS:
                result["rating"] = key
                result["rfq_reasons"].append(f"비표준 압력등급: {key}")
    # Fallback
    if not result["rating"]:
        for alias, val in RATING_MAP.items():
            if alias in q:
                result["rating"] = val
                break

    # 3. Size — try inch first (12", 12인치), then mm (300A)
    m = re.search(r'(\d+(?:\s*1/2)?)\s*[""\'인치]', raw)
    if m:
        inch_str = m.group(1).strip()
        if "1/2" in inch_str:
            inch_val = int(inch_str.split()[0]) + 0.5
        else:
            inch_val = float(inch_str)
        result["size_mm"] = SIZE_INCH_TO_MM.get(inch_val)

    if not result["size_mm"]:
        m = re.search(r"\b(\d{2,3})\s*A\b", q)
        if m:
            mm = int(m.group(1))
            if mm in SIZE_INCH_TO_MM.values():
                result["size_mm"] = mm

    # Bare inch number without quote mark (e.g., "12" in context)
    if not result["size_mm"]:
        m = re.search(r'\b(\d{1,2})\s*"', q)
        if m:
            inch_val = float(m.group(1))
            result["size_mm"] = SIZE_INCH_TO_MM.get(inch_val)

    # 4. End connection
    for ec in ["BW", "BUTT WELD", "BUTTWELD"]:
        if ec in q:
            result["end_connection"] = "BW"
            result["rfq_reasons"].append("비표준 연결방식: BW")
            break
    if not result["end_connection"]:
        for ec in SPECIAL_END_CONN:
            if re.search(r"\b" + ec + r"\b", q):
                result["end_connection"] = ec
                result["rfq_reasons"].append(f"비표준 연결방식: {ec}")
                break
    if not result["end_connection"]:
        if "RF" in q or "FLGD" in q or "FLANGE" in q:
            result["end_connection"] = "FLGD RF"
        else:
            result["end_connection"] = "FLGD RF"  # default

    # 5. Trim
    m = re.search(r"13CR\+HF|HF|STL|304SS|13CR", q)
    if m:
        result["trim"] = m.group(0)
        if "HF" in result["trim"]:
            result["rfq_reasons"].append("특수 트림: HF (하드페이싱)")
        if "STL" in result["trim"]:
            result["rfq_reasons"].append("특수 트림: STL (스텔라이트)")

    # 6. Operation
    if "GEAR" in q:
        result["operation"] = "GEAR"
        if result["size_mm"] and result["size_mm"] < 150:
            result["rfq_reasons"].append(f"소구경 기어: {result['size_mm']}A")

    return result


def lookup_price(spec: dict, discount_rate: float = 0.40) -> dict:
    """Look up price from DB. Returns {status, unit_price, message}."""
    if not spec["body_type"]:
        return {"status": "MISSING_INFO", "unit_price": None, "message": "제품군 미입력"}
    if not spec["size_mm"]:
        return {"status": "MISSING_INFO", "unit_price": None, "message": "사이즈 미입력"}

    if spec["rfq_reasons"]:
        return {
            "status": "NEEDS_MAKER_QUOTE",
            "unit_price": None,
            "message": " / ".join(spec["rfq_reasons"]),
        }

    rating = spec["rating"] or "10K"

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        """SELECT pli.unit_price FROM price_list_items pli
           JOIN products p ON pli.product_id = p.id
           WHERE p.body_type = ? AND p.rating = ? AND p.size_mm = ?
             AND pli.discount_rate = ?""",
        (spec["body_type"], rating, spec["size_mm"], discount_rate),
    ).fetchone()
    conn.close()

    if row:
        return {"status": "FOUND", "unit_price": row["unit_price"], "message": None}
    return {"status": "NOT_FOUND", "unit_price": None, "message": f"DB에 없음: {spec['body_type']} {rating} {spec['size_mm']}A"}


def create_template():
    """Create the input Excel file if it doesn't exist."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "견적"

    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [
        ("A", "사양 입력", 45),
        ("B", "할인율", 10),
        ("C", "제품군", 14),
        ("D", "압력", 8),
        ("E", "사이즈", 8),
        ("F", "단가 (KRW)", 15),
        ("G", "상태", 22),
        ("H", "비고", 35),
    ]

    for col, label, width in headers:
        cell = ws[f"{col}1"]
        cell.value = label
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        ws.column_dimensions[col].width = width

    # Sample inputs
    samples = [
        'GATE WCB/13CR RF 10K 80A',
        'GLOBE 20K 100A FLGD RF',
        'SW-CHECK 10K 200A',
        'Y-STRAINER 10K 3"',
        'GATE WCB/13CR+HF 300# BW GEAR 12"',
        'GATE 10K 50A',
        'GLOBE 150# 6"',
        '게이트밸브 80A 10K',
    ]

    for i, s in enumerate(samples, start=2):
        ws[f"A{i}"] = s
        ws[f"A{i}"].border = thin
        ws[f"B{i}"] = 0.40
        ws[f"B{i}"].border = thin
        ws[f"B{i}"].number_format = "0%"
        for col in "CDEFGH":
            ws[f"{col}{i}"].border = thin

    # Empty rows for user input
    for i in range(len(samples) + 2, len(samples) + 22):
        ws[f"A{i}"].border = thin
        ws[f"B{i}"] = 0.40
        ws[f"B{i}"].border = thin
        ws[f"B{i}"].number_format = "0%"
        for col in "CDEFGH":
            ws[f"{col}{i}"].border = thin

    wb.save(EXCEL_PATH)
    print(f"Created: {EXCEL_PATH}")


def process():
    """Read Excel, parse specs, lookup prices, write results back."""
    if not os.path.exists(EXCEL_PATH):
        create_template()
        print("Template created with samples. Run again after adding your specs.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")

    processed = 0
    for row in range(2, ws.max_row + 1):
        raw = ws[f"A{row}"].value
        if not raw or not str(raw).strip():
            continue

        raw = str(raw).strip()
        discount = ws[f"B{row}"].value
        if discount is None:
            discount = 0.40
        discount = float(discount)

        spec = parse_spec(raw)
        result = lookup_price(spec, discount)

        # Write parsed fields
        ws[f"C{row}"] = spec["body_type"] or ""
        ws[f"D{row}"] = spec["rating"] or ""
        ws[f"E{row}"] = f"{spec['size_mm']}A" if spec["size_mm"] else ""

        # Write result
        if result["unit_price"]:
            ws[f"F{row}"] = result["unit_price"]
            ws[f"F{row}"].number_format = "#,##0"
        else:
            ws[f"F{row}"] = ""

        ws[f"G{row}"] = result["status"]
        ws[f"H{row}"] = result["message"] or ""

        # Color
        if result["status"] == "FOUND":
            ws[f"G{row}"].fill = green
        elif result["status"] == "NEEDS_MAKER_QUOTE":
            ws[f"G{row}"].fill = yellow
        else:
            ws[f"G{row}"].fill = red

        processed += 1
        status_icon = {"FOUND": "O", "NEEDS_MAKER_QUOTE": "!", "MISSING_INFO": "X", "NOT_FOUND": "X"}
        price_str = f"{result['unit_price']:,}" if result["unit_price"] else "-"
        print(f"  [{status_icon.get(result['status'], '?')}] {raw[:40]:<40} → {price_str}")

    wb.save(EXCEL_PATH)
    print(f"\nDone. {processed} rows processed. Open: {EXCEL_PATH}")


if __name__ == "__main__":
    print("MSM 협가표 단가 자동 조회")
    print("=" * 50)
    process()
