"""
Migrate all Excel data into SQLite.
  - 수주대장 (한국밸브 서울 영업소 수주대장 26년.xls)
  - 발주목록 (한국밸브 발주목록.xlsx)
  - 매입 내역 (한국밸브 매입.xlsx)
"""
import sqlite3
import os
import sys
import re
from datetime import datetime, date

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "msm.sqlite")
SCHEMA_PATH = os.path.join(BASE_DIR, "db", "schema.sql")
EXCEL_DIR = os.path.join(os.path.dirname(BASE_DIR), "files for explnation")

os.makedirs(DATA_DIR, exist_ok=True)


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    with open(SCHEMA_PATH) as f:
        conn.executescript(f.read())
    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def xldate_to_date(xldate, datemode=0):
    """Convert Excel serial date to Python date. Returns None on failure."""
    if xldate is None or xldate == '' or xldate == 0:
        return None
    try:
        import xlrd
        dt = xlrd.xldate_as_tuple(xldate, datemode)
        return date(dt[0], dt[1], dt[2]).isoformat()
    except Exception:
        return None


def openpyxl_date(val):
    """Convert openpyxl date value to ISO string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)) and val > 40000:
        try:
            import xlrd
            dt = xlrd.xldate_as_tuple(val, 0)
            return date(dt[0], dt[1], dt[2]).isoformat()
        except Exception:
            pass
    return None


def safe_int(val):
    """Convert to int, return None if not possible."""
    if val is None or val == '' or val == ' ':
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val):
    if val is None or val == '' or val == ' ':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# 1. 수주대장 Migration
# ---------------------------------------------------------------------------

def migrate_orders(conn):
    """Migrate 수주대장 from .xls file."""
    import xlrd

    fpath = os.path.join(EXCEL_DIR, "한국밸브 서울 영업소 수주대장 26년.xls")
    if not os.path.exists(fpath):
        print(f"  SKIP: {fpath} not found")
        return 0

    wb = xlrd.open_workbook(fpath)
    total = 0

    # Monthly sheets to process
    monthly_sheets = []
    for sn in wb.sheet_names():
        # Match patterns like "2023년 3월", "1월", "2월", "3월", "4월", "5월"
        m = re.match(r'(\d{4})년\s*(\d{1,2})월', sn)
        if m:
            year, month = int(m.group(1)), int(m.group(2))
            monthly_sheets.append((sn, year, month))
        elif re.match(r'(\d{1,2})월', sn.strip()):
            # Standalone month sheets are 2026
            m2 = re.match(r'(\d{1,2})월', sn.strip())
            monthly_sheets.append((sn, 2026, int(m2.group(1))))

    print(f"  Found {len(monthly_sheets)} monthly sheets")

    for sheet_name, year, month in monthly_sheets:
        ws = wb.sheet_by_name(sheet_name)
        year_month = f"{year}-{month:02d}"

        # Find header row (contains '발주NO' or 'ITEM')
        header_row = None
        for r in range(min(10, ws.nrows)):
            for c in range(ws.ncols):
                v = ws.cell(r, c).value
                if v and '발주NO' in str(v):
                    header_row = r
                    break
                if v and 'ITEM' in str(v):
                    header_row = r
                    break
            if header_row is not None:
                break

        if header_row is None:
            continue

        # Build column mapping from header
        col_map = {}
        for c in range(ws.ncols):
            v = str(ws.cell(header_row, c).value).strip()
            if '발주NO' in v or v == '발주NO':
                col_map['order_no'] = c
            elif 'ITEM' in v:
                col_map['item'] = c
            elif '업 체 명' in v or '업체명' == v:
                col_map['customer'] = c
            elif '단가' in v:
                col_map['unit_price'] = c
            elif '수  량' in v or '수량' in v:
                col_map['qty'] = c
            elif v == '금액' and 'amount' not in col_map:
                col_map['amount'] = c
            elif '수주일자' in v:
                col_map['order_date'] = c
            elif '납기일자' in v and 'delivery_due' not in col_map:
                col_map['delivery_due'] = c
            elif '납품일자' in v:
                col_map['delivery_date'] = c
            elif '납품금액' in v:
                col_map['delivery_amount'] = c
            elif '매출일자' in v:
                col_map['sales_date'] = c
            elif '매출금액' in v:
                col_map['sales_amount'] = c
            elif v == 'REMARK' and 'remark' not in col_map:
                col_map['remark'] = c
            elif '수금일자' in v:
                col_map['collection_date'] = c
            elif '재고판매' in v:
                col_map['stock_amount'] = c

        # Also look for 매입 columns (usually starting at col 15+)
        for c in range(ws.ncols):
            v = str(ws.cell(header_row, c).value).strip()
            if c >= 14:
                if '업체명' in v:
                    col_map['supplier'] = c
                elif '발주일자' in v:
                    col_map['purchase_date'] = c
                elif '금액' in v:
                    col_map['purchase_amount'] = c
                elif '납기일자' in v:
                    col_map['purchase_due'] = c
                elif '납품사양' in v:
                    col_map['purchase_spec'] = c
                elif '계산서' in v:
                    col_map['invoice_date'] = c
                elif '원가율' in v:
                    col_map['cost_ratio'] = c

        # Check for 특이사항 column (col 2 in newer sheets)
        for c in range(ws.ncols):
            v = str(ws.cell(header_row, c).value).strip()
            if '특이사항' in v:
                col_map['memo_col'] = c

        def cell(r, col_name):
            c = col_map.get(col_name)
            if c is None:
                return None
            if r >= ws.nrows:
                return None
            return ws.cell(r, c).value

        # Parse data rows
        data_start = header_row + 1
        current_order = None
        order_items = []

        for r in range(data_start, ws.nrows):
            no_val = ws.cell(r, 0).value
            # Check if this is a TOTAL row
            customer_raw = cell(r, 'customer')
            if customer_raw and 'TOTAL' in str(customer_raw):
                # Save current order
                if current_order:
                    oid = save_order(conn, current_order, order_items)
                    if oid:
                        total += 1
                    current_order = None
                    order_items = []
                continue

            # New order starts when col 0 has a number
            if isinstance(no_val, (int, float)) and no_val > 0 and no_val != ' ':
                # Save previous order
                if current_order:
                    oid = save_order(conn, current_order, order_items)
                    if oid:
                        total += 1

                # Determine discount_rate or memo from col 2
                discount_rate = None
                memo = None
                memo_val = cell(r, 'memo_col') if 'memo_col' in col_map else ws.cell(r, 2).value if ws.ncols > 2 else None
                if isinstance(memo_val, (int, float)) and 0 < memo_val < 1:
                    discount_rate = memo_val
                elif memo_val and str(memo_val).strip():
                    s = str(memo_val).strip()
                    # Check if it's a discount like "0.47"
                    try:
                        fv = float(s)
                        if 0 < fv < 1:
                            discount_rate = fv
                        else:
                            memo = s
                    except ValueError:
                        memo = s

                current_order = {
                    'order_no': safe_str(cell(r, 'order_no')),
                    'order_seq': safe_int(no_val),
                    'year_month': year_month,
                    'customer_name': safe_str(cell(r, 'customer')),
                    'discount_rate': discount_rate,
                    'memo': memo,
                    'order_date': xldate_to_date(cell(r, 'order_date'), wb.datemode),
                    'total_amount': safe_int(cell(r, 'amount')),
                    'stock_amount': safe_int(cell(r, 'stock_amount')),
                    'delivery_due': xldate_to_date(cell(r, 'delivery_due'), wb.datemode),
                    'delivery_date': xldate_to_date(cell(r, 'delivery_date'), wb.datemode),
                    'delivery_amount': safe_int(cell(r, 'delivery_amount')),
                    'sales_date': xldate_to_date(cell(r, 'sales_date'), wb.datemode),
                    'sales_amount': safe_int(cell(r, 'sales_amount')),
                    'remark': safe_str(cell(r, 'remark')),
                    'collection_date': safe_str(cell(r, 'collection_date')),
                    'supplier_name': safe_str(cell(r, 'supplier')),
                    'purchase_date': xldate_to_date(cell(r, 'purchase_date'), wb.datemode),
                    'purchase_amount': safe_int(cell(r, 'purchase_amount')),
                    'purchase_due': xldate_to_date(cell(r, 'purchase_due'), wb.datemode),
                    'purchase_spec': safe_str(cell(r, 'purchase_spec')),
                    'invoice_date': xldate_to_date(cell(r, 'invoice_date'), wb.datemode),
                    'source_file': os.path.basename(fpath),
                    'source_sheet': sheet_name,
                }
                order_items = []

            # Collect item rows
            item_desc = safe_str(cell(r, 'item'))
            if item_desc and current_order:
                order_items.append({
                    'item_desc': item_desc,
                    'unit_price': safe_int(cell(r, 'unit_price')),
                    'quantity': safe_int(cell(r, 'qty')),
                    'amount': safe_int(cell(r, 'amount')),
                    'purchase_spec': safe_str(cell(r, 'purchase_spec')),
                })

            # Accumulate amounts from continuation rows
            if current_order and not isinstance(no_val, (int, float)):
                amt = safe_int(cell(r, 'amount'))
                if amt and current_order['total_amount']:
                    current_order['total_amount'] += amt
                elif amt:
                    current_order['total_amount'] = amt

                # Pick up supplier from continuation rows
                sup = safe_str(cell(r, 'supplier'))
                if sup and not current_order['supplier_name']:
                    current_order['supplier_name'] = sup

                # Pick up additional memo from col 2
                if 'memo_col' in col_map:
                    extra = safe_str(ws.cell(r, col_map['memo_col']).value)
                else:
                    extra = safe_str(ws.cell(r, 2).value) if ws.ncols > 2 else None
                if extra and extra != 'TOTAL':
                    if isinstance(extra, str) and not re.match(r'^[\d.]+$', extra):
                        if current_order['memo']:
                            current_order['memo'] += '; ' + extra
                        else:
                            current_order['memo'] = extra

        # Save last order
        if current_order:
            oid = save_order(conn, current_order, order_items)
            if oid:
                total += 1

        print(f"    {sheet_name}: migrated orders for {year_month}")

    conn.commit()
    return total


def save_order(conn, order, items):
    """Insert order + items, return order id."""
    if not order.get('customer_name') and not order.get('order_no'):
        return None

    # Handle collection_date which might be text like "입금완결"
    collection_note = None
    collection_date = order.get('collection_date')
    if collection_date and not re.match(r'\d{4}-\d{2}-\d{2}', str(collection_date)):
        collection_note = str(collection_date)
        collection_date = None

    cur = conn.execute("""
        INSERT INTO orders (
            order_no, order_seq, year_month, customer_name, discount_rate, memo,
            order_date, total_amount, stock_amount, delivery_due, delivery_date,
            delivery_amount, sales_date, sales_amount, remark,
            collection_date, collection_note,
            supplier_name, purchase_date, purchase_amount, purchase_due,
            purchase_spec, invoice_date, source_file, source_sheet
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        order.get('order_no'), order.get('order_seq'), order.get('year_month'),
        order.get('customer_name'), order.get('discount_rate'), order.get('memo'),
        order.get('order_date'), order.get('total_amount'), order.get('stock_amount'),
        order.get('delivery_due'), order.get('delivery_date'),
        order.get('delivery_amount'), order.get('sales_date'), order.get('sales_amount'),
        order.get('remark'), collection_date, collection_note,
        order.get('supplier_name'), order.get('purchase_date'),
        order.get('purchase_amount'), order.get('purchase_due'),
        order.get('purchase_spec'), order.get('invoice_date'),
        order.get('source_file'), order.get('source_sheet'),
    ))
    order_id = cur.lastrowid

    for item in items:
        if not item.get('item_desc'):
            continue
        conn.execute("""
            INSERT INTO order_items (order_id, item_desc, unit_price, quantity, amount, purchase_spec)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            order_id, item['item_desc'], item.get('unit_price'),
            item.get('quantity'), item.get('amount'), item.get('purchase_spec'),
        ))

    return order_id


# ---------------------------------------------------------------------------
# 2. 발주목록 Migration
# ---------------------------------------------------------------------------

def migrate_purchase_orders(conn):
    """Migrate 발주목록 from .xlsx file."""
    import openpyxl

    fpath = os.path.join(EXCEL_DIR, "한국밸브 발주목록.xlsx")
    if not os.path.exists(fpath):
        print(f"  SKIP: {fpath} not found")
        return 0

    wb = openpyxl.load_workbook(fpath, data_only=True)
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Determine year from sheet name
        year = None
        if sheet_name.isdigit() and len(sheet_name) == 4:
            year = int(sheet_name)
        elif re.match(r'한국밸브\d{4}', sheet_name):
            year = int(re.search(r'(\d{4})', sheet_name).group(1))
        else:
            # Try to extract year
            m = re.search(r'(\d{4})', sheet_name)
            if m:
                year = int(m.group(1))
            else:
                continue

        # Find header row
        header_row = None
        col_map = {}
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and ('발주서 No' in str(v) or '주문번호' in str(v)):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            continue

        # Build column mapping
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(header_row, c).value or '').strip()
            if '발주서 No' in v or '주문번호' in v:
                col_map['po_number'] = c
            elif v == '매입처' or '거래처' in v:
                col_map['supplier'] = c
            elif '발주품목' in v or 'DESCRIPTION' in v:
                col_map['item'] = c
            elif "Q'TY" in v or 'QTY' in v:
                col_map['qty'] = c
            elif '금액' in v and 'amount' not in col_map:
                col_map['amount'] = c
            elif '발주일자' in v:
                col_map['order_date'] = c
            elif '납품일자' in v or '납기' in v:
                col_map['delivery_due'] = c
            elif '출고여부' in v:
                col_map['shipped'] = c
            elif '비고' in v and 'remark' not in col_map:
                col_map['remark'] = c

        def cell(r, col_name):
            c = col_map.get(col_name)
            if c is None:
                return None
            return ws.cell(r, c).value

        # Parse data rows
        for r in range(header_row + 1, ws.max_row + 1):
            po_number = safe_str(cell(r, 'po_number'))
            item_desc = safe_str(cell(r, 'item'))
            amount = safe_int(cell(r, 'amount'))

            # Skip empty or summary rows
            if not po_number and not item_desc and not amount:
                continue
            if item_desc and 'TOTAL' in str(item_desc):
                continue

            # Handle delivery_due which can be date or text
            delivery_due_raw = cell(r, 'delivery_due')
            delivery_due = None
            if delivery_due_raw:
                d = openpyxl_date(delivery_due_raw)
                if d:
                    delivery_due = d
                else:
                    delivery_due = safe_str(delivery_due_raw)

            order_date = openpyxl_date(cell(r, 'order_date'))

            conn.execute("""
                INSERT INTO purchase_orders (
                    po_number, year, supplier_name, item_desc, quantity, amount,
                    order_date, delivery_due, shipped, customer_name, remark,
                    source_file, source_sheet
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                po_number, year, safe_str(cell(r, 'supplier')),
                item_desc, safe_int(cell(r, 'qty')), amount,
                order_date, delivery_due,
                safe_str(cell(r, 'shipped')), None,
                safe_str(cell(r, 'remark')),
                os.path.basename(fpath), sheet_name,
            ))
            total += 1

        print(f"    {sheet_name}: migrated POs for year {year}")

    conn.commit()
    return total


# ---------------------------------------------------------------------------
# 3. 매입 내역 Migration
# ---------------------------------------------------------------------------

def migrate_purchases(conn):
    """Migrate 매입 from .xlsx file."""
    import openpyxl

    fpath = os.path.join(EXCEL_DIR, "한국밸브 매입.xlsx")
    if not os.path.exists(fpath):
        print(f"  SKIP: {fpath} not found")
        return 0

    wb = openpyxl.load_workbook(fpath, data_only=True)
    total = 0

    for sheet_name in wb.sheetnames:
        # Parse year-month from sheet name like "2411", "2501", "2601"
        if not re.match(r'\d{4}', sheet_name):
            continue

        yy = int(sheet_name[:2])
        mm = int(sheet_name[2:])
        year = 2000 + yy
        year_month = f"{year}-{mm:02d}"

        ws = wb[sheet_name]

        # Find header row (contains '거래명세표')
        header_row = None
        col_map = {}
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and '거래명세표' in str(v):
                    header_row = r
                    break
            if header_row:
                break

        if not header_row:
            continue

        # Build column mapping
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(header_row, c).value or '').strip()
            if '거래명세표' in v:
                col_map['invoice_no'] = c
            elif '수주 No' in v or '수주No' in v:
                col_map['order_ref'] = c
            elif '발주No' in v:
                col_map['po_ref'] = c
            elif '발주금액' in v:
                col_map['po_amount'] = c
            elif '매입액' in v:
                col_map['purchase_amount'] = c
            elif '구분' in v:
                col_map['category'] = c
            elif '비고' in v:
                col_map['memo'] = c
            elif 'MSM' in v and '매출' in v:
                col_map['sales_amount'] = c
            elif '매출이익' in v or '50%' in v:
                col_map['profit_half'] = c

        def cell(r, col_name):
            c = col_map.get(col_name)
            if c is None:
                return None
            return ws.cell(r, c).value

        # Parse data rows
        for r in range(header_row + 1, ws.max_row + 1):
            invoice_no = cell(r, 'invoice_no')
            po_ref = safe_str(cell(r, 'po_ref'))
            purchase_amount = safe_int(cell(r, 'purchase_amount'))

            # Skip empty or summary rows
            if not invoice_no and not po_ref and not purchase_amount:
                continue
            category = safe_str(cell(r, 'category'))
            if category and '합계' in str(category):
                continue
            if po_ref and '합계' in str(po_ref):
                continue
            if category and 'V.A.T' in str(category):
                continue
            if category and ('PJT' in str(category) or 'STOCK' == str(category).strip()):
                # These are summary lines, skip
                if not invoice_no:
                    continue

            conn.execute("""
                INSERT INTO purchases (
                    year_month, invoice_no, order_ref, po_ref,
                    po_amount, purchase_amount, category, memo,
                    sales_amount, profit_half, source_file, source_sheet
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                year_month, safe_str(invoice_no), safe_str(cell(r, 'order_ref')),
                po_ref, safe_int(cell(r, 'po_amount')), purchase_amount,
                category, safe_str(cell(r, 'memo')),
                safe_int(cell(r, 'sales_amount')), safe_int(cell(r, 'profit_half')),
                os.path.basename(fpath), sheet_name,
            ))
            total += 1

        print(f"    {sheet_name}: migrated purchases for {year_month}")

    conn.commit()
    return total


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("MSM Data Migration")
    print("=" * 60)
    print(f"DB: {DB_PATH}")
    print(f"Excel source: {EXCEL_DIR}")
    print()

    if not os.path.isdir(EXCEL_DIR):
        print(f"ERROR: Excel directory not found: {EXCEL_DIR}")
        sys.exit(1)

    conn = init_db()

    # Clear existing data (re-runnable)
    for table in ['order_items', 'orders', 'purchase_orders', 'purchases']:
        conn.execute(f"DELETE FROM {table}")
    conn.commit()

    print("[1/3] Migrating 수주대장...")
    n_orders = migrate_orders(conn)
    print(f"  -> {n_orders} orders migrated\n")

    print("[2/3] Migrating 발주목록...")
    n_po = migrate_purchase_orders(conn)
    print(f"  -> {n_po} purchase orders migrated\n")

    print("[3/3] Migrating 매입 내역...")
    n_purchases = migrate_purchases(conn)
    print(f"  -> {n_purchases} purchase records migrated\n")

    # Summary
    print("=" * 60)
    print("Migration Summary")
    print("=" * 60)
    for table in ['orders', 'order_items', 'purchase_orders', 'purchases',
                   'customers', 'suppliers', 'inventory']:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"  {table:25s} {count:>6,d} rows")

    conn.close()
    print(f"\nDone. Database at: {DB_PATH}")


if __name__ == "__main__":
    main()
