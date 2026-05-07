"""Generate the Excel quote template (.xlsx) with headers and sample data."""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "견적 조회"

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
input_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
output_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Title
ws.merge_cells("B1:H1")
ws["B1"] = "MSM 협가표 단가 조회"
ws["B1"].font = Font(bold=True, size=16)
ws["B1"].alignment = Alignment(horizontal="center")

# Headers (row 2)
headers = [
    ("B", "BODY TYPE", 15, input_fill),
    ("C", "RATING", 12, input_fill),
    ("D", "SIZE", 10, input_fill),
    ("E", "END CONN", 12, input_fill),
    ("F", "DISCOUNT", 12, input_fill),
    ("G", "UNIT PRICE", 15, output_fill),
    ("H", "STATUS", 20, output_fill),
]

for col, label, width, fill in headers:
    cell = ws[f"{col}2"]
    cell.value = label
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin
    ws.column_dimensions[col].width = width

# Sample data rows
samples = [
    ("GATE", "10K", "50A", "FLGD RF", "0.40"),
    ("GATE", "10K", "80A", "FLGD RF", "0.40"),
    ("GLOBE", "20K", "100A", "FLGD RF", "0.45"),
    ("SW-CHECK", "10K", "200A", "FLGD RF", "0.50"),
    ("Y-STRAINER", "10K", "80A", "FLGD RF", "0.40"),
    ("GATE", "600#", "80A", "FLGD RF", "0.40"),
    ("GATE", "10K", "80A", "SW", "0.40"),
    ("", "", "", "", ""),
    ("", "", "", "", ""),
    ("", "", "", "", ""),
]

for i, (body, rating, size, end, disc) in enumerate(samples):
    r = i + 3
    for col_idx, val in enumerate([body, rating, size, end, disc], start=2):
        cell = ws.cell(row=r, column=col_idx)
        cell.value = val
        cell.border = thin
        if col_idx <= 6 and val:
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    # Output columns
    ws.cell(row=r, column=7).border = thin
    ws.cell(row=r, column=7).number_format = "#,##0"
    ws.cell(row=r, column=8).border = thin

# Instructions
ws["B14"] = "사용법:"
ws["B14"].font = Font(bold=True)
ws["B15"] = "1. 서버 실행: python3 -m uvicorn api.main:app --port 8000"
ws["B16"] = "2. B~F열에 사양 입력"
ws["B17"] = "3. VBA 매크로 LookupPrice 실행 (또는 LookupAll로 전체 조회)"
ws["B18"] = ""
ws["B19"] = "BODY TYPE: GATE, GLOBE, SW-CHECK, Y-STRAINER"
ws["B20"] = "RATING: 10K, 20K (or 150#, 300#)"
ws["B21"] = "DISCOUNT: 0.0 (정가) ~ 0.56 (56%)"

out_path = os.path.join(os.path.dirname(__file__), "..", "excel", "quote_template.xlsx")
wb.save(out_path)
print(f"Template saved to: {os.path.abspath(out_path)}")
