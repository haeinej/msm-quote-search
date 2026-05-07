"""
Build a self-contained Excel file with:
- Sheet "조회": dropdown selectors → price auto-fills via formulas
- Sheet "DB": flat price table used by INDEX/MATCH
No Python or server needed after this file is generated.
"""
import os
import sys
import sqlite3

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(DIR, "..", "data", "msm.sqlite")
OUT_PATH = os.path.join(DIR, "..", "MSM_협가표_조회.xlsx")

# --- Load all prices from DB ---
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
rows = conn.execute("""
    SELECT p.body_type, p.rating, p.size_mm, pli.discount_rate, pli.unit_price
    FROM price_list_items pli
    JOIN products p ON pli.product_id = p.id
    ORDER BY p.body_type, p.rating, p.size_mm, pli.discount_rate
""").fetchall()
conn.close()

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: DB (flat lookup table)
# ============================================================
ws_db = wb.active
ws_db.title = "DB"

thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF", size=10)

db_headers = ["제품군", "압력등급", "사이즈", "할인율", "단가", "KEY"]
for i, h in enumerate(db_headers, 1):
    c = ws_db.cell(row=1, column=i, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.border = thin
    c.alignment = Alignment(horizontal="center")

ws_db.column_dimensions["A"].width = 14
ws_db.column_dimensions["B"].width = 10
ws_db.column_dimensions["C"].width = 8
ws_db.column_dimensions["D"].width = 10
ws_db.column_dimensions["E"].width = 14
ws_db.column_dimensions["F"].width = 30

for idx, r in enumerate(rows, start=2):
    discount_pct = int(r["discount_rate"] * 100)
    discount_label = f"-{discount_pct}%" if discount_pct > 0 else "0%"
    key = f"{r['body_type']}_{r['rating']}_{r['size_mm']}A_{discount_label}"

    ws_db.cell(row=idx, column=1, value=r["body_type"]).border = thin
    ws_db.cell(row=idx, column=2, value=r["rating"]).border = thin
    ws_db.cell(row=idx, column=3, value=f"{r['size_mm']}A").border = thin
    ws_db.cell(row=idx, column=4, value=discount_label).border = thin
    c = ws_db.cell(row=idx, column=5, value=r["unit_price"])
    c.border = thin
    c.number_format = "#,##0"
    ws_db.cell(row=idx, column=6, value=key).border = thin

db_row_count = len(rows) + 1  # including header

# ============================================================
# Sheet 2: 조회 (search interface)
# ============================================================
ws = wb.create_sheet("조회", 0)  # make it the first sheet

# --- Title ---
ws.merge_cells("B1:G1")
ws["B1"] = "MSM 협가표 단가 조회"
ws["B1"].font = Font(bold=True, size=18, color="1F4E79")
ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

ws.merge_cells("B2:G2")
ws["B2"] = "한국밸브 CAST CARBON STEEL VALVE [SCPH2/WCB] — 2022.01.01"
ws["B2"].font = Font(size=10, color="666666")
ws["B2"].alignment = Alignment(horizontal="center")

# --- Input area ---
input_fill = PatternFill("solid", fgColor="D6E4F0")
result_fill = PatternFill("solid", fgColor="E2EFDA")
label_font = Font(bold=True, size=11, color="1F4E79")

# Unique values for dropdowns
body_types = sorted(set(r["body_type"] for r in rows))
ratings = sorted(set(r["rating"] for r in rows))
sizes_raw = sorted(set(r["size_mm"] for r in rows))
sizes = [f"{s}A" for s in sizes_raw]
discount_rates = sorted(set(r["discount_rate"] for r in rows))
discount_labels = []
for d in discount_rates:
    pct = int(d * 100)
    discount_labels.append(f"-{pct}%" if pct > 0 else "0%")

# Headers row 4
labels = [
    ("B", "제품군"),
    ("C", "압력등급"),
    ("D", "사이즈"),
    ("E", "할인율"),
    ("F", "단가 (KRW)"),
    ("G", "상태"),
]
widths = {"B": 16, "C": 12, "D": 10, "E": 12, "F": 18, "G": 16}

for col, label in labels:
    c = ws[f"{col}4"]
    c.value = label
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin
    ws.column_dimensions[col].width = widths[col]

ws.column_dimensions["A"].width = 3
ws.column_dimensions["H"].width = 3

# --- Data validation (dropdowns) ---
dv_body = DataValidation(type="list", formula1=f'"{",".join(body_types)}"', allow_blank=True)
dv_body.error = "제품군을 선택하세요"
dv_body.prompt = "GATE, GLOBE, SW-CHECK, Y-STRAINER"
dv_body.promptTitle = "제품군"

dv_rating = DataValidation(type="list", formula1=f'"{",".join(ratings)}"', allow_blank=True)
dv_pressure = dv_rating

dv_size = DataValidation(type="list", formula1=f'"{",".join(sizes)}"', allow_blank=True)

dv_discount = DataValidation(type="list", formula1=f'"{",".join(discount_labels)}"', allow_blank=True)

ws.add_data_validation(dv_body)
ws.add_data_validation(dv_rating)
ws.add_data_validation(dv_size)
ws.add_data_validation(dv_discount)

# --- Rows 5-24: input rows with formulas ---
NUM_ROWS = 20

for i in range(NUM_ROWS):
    r = 5 + i

    # Input cells with dropdowns
    for col in ["B", "C", "D", "E"]:
        c = ws[f"{col}{r}"]
        c.border = thin
        c.fill = input_fill
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=11)

    dv_body.add(ws[f"B{r}"])
    dv_rating.add(ws[f"C{r}"])
    dv_size.add(ws[f"D{r}"])
    dv_discount.add(ws[f"E{r}"])

    # KEY formula (hidden helper in column I)
    ws[f"I{r}"] = f'=IF(B{r}="","",B{r}&"_"&C{r}&"_"&D{r}&"_"&E{r})'
    ws[f"I{r}"].font = Font(color="FFFFFF", size=1)

    # Price lookup: INDEX/MATCH on the KEY column in DB sheet
    price_formula = f'=IF(I{r}="","",IFERROR(INDEX(DB!$E$2:$E${db_row_count},MATCH(I{r},DB!$F$2:$F${db_row_count},0)),""))'
    c = ws[f"F{r}"]
    c.value = price_formula
    c.border = thin
    c.fill = result_fill
    c.number_format = "#,##0"
    c.font = Font(size=12, bold=True, color="1A73E8")
    c.alignment = Alignment(horizontal="right")

    # Status formula
    status_formula = (
        f'=IF(B{r}="","",'
        f'IF(F{r}="","NOT_FOUND",'
        f'IF(F{r}>0,"FOUND","NOT_FOUND")))'
    )
    c = ws[f"G{r}"]
    c.value = status_formula
    c.border = thin
    c.alignment = Alignment(horizontal="center")
    c.font = Font(size=10)

    # Default discount
    ws[f"E{r}"] = "-40%"

# --- Instructions ---
r_start = 5 + NUM_ROWS + 2
ws[f"B{r_start}"] = "사용법"
ws[f"B{r_start}"].font = Font(bold=True, size=12, color="1F4E79")
ws[f"B{r_start + 1}"] = "1. 제품군, 압력등급, 사이즈를 드롭다운에서 선택"
ws[f"B{r_start + 2}"] = "2. 할인율 선택 (기본 -40%)"
ws[f"B{r_start + 3}"] = "3. 단가가 자동으로 표시됩니다"
ws[f"B{r_start + 5}"] = "할인율 범위: 0% (정가), -40% ~ -56%"
ws[f"B{r_start + 6}"] = "DB 시트에 전체 가격 데이터 (1,512건) 포함"

for i in range(7):
    ws[f"B{r_start + 1 + i}"].font = Font(size=10, color="444444")

# Hide helper column I and DB sheet key column F
ws.column_dimensions["I"].hidden = True

# Set 조회 as the active sheet
wb.active = 0

wb.save(OUT_PATH)
print(f"Done: {OUT_PATH}")
print(f"  {len(rows)} prices loaded across {len(discount_labels)} discount tiers")
print(f"  {NUM_ROWS} input rows with dropdown + auto-lookup")
