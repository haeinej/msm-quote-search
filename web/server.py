"""
MSM Valve Management System — FastAPI Web Server
Uses Supabase REST API if configured, else falls back to SQLite.
"""
import os
import sys
import io
import secrets
from datetime import date, datetime
from typing import Optional

from fastapi import FastAPI, Query, Request, UploadFile, File, Body, Cookie, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
import re as _re

load_dotenv()

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from parser import parse_query
from db.connection import is_supabase, get_supabase, get_sqlite, get_init_error

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")

# ============================================================
# Auth
# ============================================================
INTERNAL_PASSWORD = os.environ.get("MSM_PASSWORD", "msm2026!")
_valid_tokens: set[str] = set()


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        if path in ("/api/login", "/login", "/api/health") or path.startswith("/static"):
            return await call_next(request)
        token = request.cookies.get("msm_token")
        if not token or token not in _valid_tokens:
            if path.startswith("/api/"):
                return JSONResponse({"detail": "Unauthorized"}, status_code=401)
            login_path = os.path.join(STATIC_DIR, "login.html")
            with open(login_path, encoding="utf-8") as f:
                return HTMLResponse(f.read())
        return await call_next(request)


app = FastAPI(title="MSM Valve Management", version="2.0.0")
app.add_middleware(AuthMiddleware)


@app.get("/api/health")
def health():
    """Health check — no auth required, useful for debugging."""
    import sys
    info = {
        "status": "ok",
        "python": sys.version[:10],
        "supabase_configured": is_supabase(),
        "supabase_url_set": bool(os.environ.get("SUPABASE_URL")),
        "supabase_key_set": bool(os.environ.get("SUPABASE_SERVICE_KEY")),
        "init_error": get_init_error(),
    }
    if is_supabase():
        try:
            r = sb().table("price_catalog").select("id", count="exact").limit(1).execute()
            info["db_connected"] = True
            info["price_catalog_count"] = r.count
        except Exception as e:
            info["db_connected"] = False
            info["db_error"] = str(e)[:500]
    return info


@app.post("/api/login")
async def login(request: Request):
    body = await request.json()
    if body.get("password") == INTERNAL_PASSWORD:
        token = secrets.token_hex(32)
        _valid_tokens.add(token)
        response = JSONResponse({"ok": True})
        response.set_cookie("msm_token", token, httponly=True, samesite="lax", max_age=86400 * 30)
        return response
    return JSONResponse({"ok": False, "error": "비밀번호가 틀렸습니다"}, status_code=403)


@app.post("/api/logout")
async def logout(msm_token: Optional[str] = Cookie(None)):
    if msm_token:
        _valid_tokens.discard(msm_token)
    response = JSONResponse({"ok": True})
    response.delete_cookie("msm_token")
    return response


@app.get("/flow", response_class=HTMLResponse)
def flow_page():
    flow_path = os.path.join(STATIC_DIR, "flow.html")
    with open(flow_path, encoding="utf-8") as f:
        return HTMLResponse(f.read())


@app.get("/api/activity-log")
def get_activity_log(limit: int = Query(100)):
    """Get recent activity log."""
    if is_supabase():
        return _sq("activity_log").select("*").order("created_at", desc=True).limit(limit).execute().data
    return []


@app.get("/api/workflow/alerts")
def get_alerts():
    """Count pending receives and invoice mismatches."""
    pending = 0
    mismatch = 0
    if is_supabase():
        r = _sq("orders").select("id", count="exact").eq("state", "발주중").execute()
        pending = r.count or 0
    return {"pending_receive": pending, "mismatch": mismatch}


# ============================================================
# Helpers
# ============================================================

def sb():
    return get_supabase()

def _sq(table):
    return sb().table(table)

def _safe_col(field: str) -> str:
    """Validate a column name to prevent SQL injection via f-string interpolation."""
    if not _re.match(r'^[a-z_]+$', field):
        raise HTTPException(400, f"Invalid field name: {field}")
    return field

def _log(order_id, action, detail=None):
    """Log an activity to the activity_log table."""
    try:
        if is_supabase():
            _sq("activity_log").insert({"order_id": order_id, "action": action, "detail": detail}).execute()
    except Exception:
        pass  # Don't fail the main operation if logging fails


# ============================================================
# API: Price Search
# ============================================================

@app.get("/api/search")
def search_price(q: str = Query(""), discount: str = Query("")):
    parsed = parse_query(q)
    if discount:
        parsed["discount_rate"] = discount

    if is_supabase():
        query = _sq("price_catalog").select("*")
        if parsed.get("product_type"):
            query = query.eq("product_type", parsed["product_type"])
        if parsed.get("pressure_class"):
            query = query.eq("pressure_class", parsed["pressure_class"])
        if parsed.get("size_a"):
            query = query.eq("size_a", parsed["size_a"])
        if parsed.get("discount_rate"):
            query = query.eq("discount_rate", parsed["discount_rate"])
        rows = query.order("product_type").order("pressure_class").order("size_a").order("discount_rate").execute().data
    else:
        conn = get_sqlite("한국밸브_협가표.db")
        conditions, params = [], []
        for key, col in [("product_type","product_type"),("pressure_class","pressure_class"),("size_a","size_a"),("discount_rate","discount_rate")]:
            if parsed.get(key):
                conditions.append(f"{col} = ?"); params.append(parsed[key])
        where = " AND ".join(conditions) if conditions else "1=1"
        rows = [dict(r) for r in conn.execute(f"SELECT * FROM price_list_items WHERE {where} ORDER BY product_type,pressure_class,size_a,discount_rate", params).fetchall()]
        conn.close()

    return {
        "parsed": {k: v for k, v in parsed.items() if k not in ("raw_query", "material") and v is not None and v != []},
        "count": len(rows),
        "status": "exact" if len(rows) == 1 else ("multiple" if rows else "none"),
        "results": rows,
    }


# ============================================================
# API: Orders (수주대장)
# ============================================================

@app.get("/api/orders/rows")
def list_order_rows(year_month: str = Query(""), q: str = Query("")):
    if is_supabase():
        oq = _sq("orders").select("*")
        if year_month: oq = oq.eq("year_month", year_month)
        if q: oq = oq.or_(f"customer_name.ilike.%{q}%,memo.ilike.%{q}%")
        orders = oq.order("year_month", desc=True).order("order_seq").execute().data
        if not orders: return []
        oids = [o["id"] for o in orders]
        all_items = []
        for i in range(0, len(oids), 50):
            all_items.extend(_sq("order_items").select("*").in_("order_id", oids[i:i+50]).order("id").execute().data)
        items_map = {}
        for it in all_items: items_map.setdefault(it["order_id"], []).append(it)
        rows = []
        for o in orders:
            oi_list = items_map.get(o["id"], [{"id":None,"item_desc":None,"unit_price":None,"quantity":None,"amount":None}])
            for it in oi_list:
                rows.append({**{k: o.get(k) for k in ["year_month","order_seq","customer_name","memo","stock_amount","order_date","delivery_due","delivery_date","delivery_amount","sales_date","sales_amount","remark","collection_note","supplier_name","purchase_date","purchase_amount","purchase_due","discount_rate"]},
                    "order_id": o["id"], "item_id": it.get("id"), "item_desc": it.get("item_desc"),
                    "unit_price": it.get("unit_price"), "quantity": it.get("quantity"), "amount": it.get("amount")})
        return rows
    else:
        conn = get_sqlite()
        conditions, params = [], []
        if year_month: conditions.append("o.year_month = ?"); params.append(year_month)
        if q: conditions.append("(o.customer_name LIKE ? OR o.memo LIKE ? OR oi.item_desc LIKE ?)"); params.extend([f"%{q}%"]*3)
        where = " AND ".join(conditions) if conditions else "1=1"
        rows = [dict(r) for r in conn.execute(f"""
            SELECT o.id as order_id, o.year_month, o.order_seq, o.customer_name, o.memo,
                oi.id as item_id, oi.item_desc, oi.unit_price, oi.quantity, oi.amount,
                o.stock_amount, o.order_date, o.delivery_due, o.delivery_date,
                o.delivery_amount, o.sales_date, o.sales_amount, o.remark,
                o.collection_note, o.supplier_name, o.purchase_date,
                o.purchase_amount, o.purchase_due, o.discount_rate
            FROM orders o LEFT JOIN order_items oi ON oi.order_id = o.id
            WHERE {where} ORDER BY o.year_month DESC, o.order_seq ASC, oi.id ASC
        """, params).fetchall()]
        conn.close()
        return rows

@app.get("/api/orders/months")
def order_months():
    if is_supabase():
        result = _sq("orders").select("year_month").execute()
        return sorted(set(r["year_month"] for r in result.data if r["year_month"]), reverse=True)
    else:
        conn = get_sqlite()
        rows = conn.execute("SELECT DISTINCT year_month FROM orders ORDER BY year_month DESC").fetchall()
        conn.close()
        return [r["year_month"] for r in rows]

class OrderRowUpdate(BaseModel):
    order_id: int
    item_id: Optional[int] = None
    field: str
    value: Optional[str] = None

@app.post("/api/orders/update-cell")
def update_order_cell(update: OrderRowUpdate):
    item_fields = {"item_desc","unit_price","quantity","amount"}
    order_fields = {"customer_name","memo","order_date","stock_amount","delivery_due","delivery_date","delivery_amount","sales_date","sales_amount","remark","collection_note","supplier_name","purchase_date","purchase_amount","purchase_due","order_seq","discount_rate"}
    field, value = update.field, update.value if update.value != "" else None
    numeric = {"unit_price","quantity","amount","stock_amount","delivery_amount","sales_amount","purchase_amount"}
    if field in numeric and value is not None:
        try: value = int(float(str(value).replace(",","")))
        except (ValueError, TypeError): pass
    if is_supabase():
        if field in item_fields and update.item_id: _sq("order_items").update({field:value}).eq("id",update.item_id).execute()
        elif field in order_fields: _sq("orders").update({field:value}).eq("id",update.order_id).execute()
        else: return {"ok":False,"error":f"Unknown field: {field}"}
    else:
        conn = get_sqlite()
        col = _safe_col(field)
        if field in item_fields and update.item_id: conn.execute(f"UPDATE order_items SET {col}=? WHERE id=?",(value,update.item_id))
        elif field in order_fields: conn.execute(f"UPDATE orders SET {col}=? WHERE id=?",(value,update.order_id))
        else: conn.close(); return {"ok":False,"error":f"Unknown field: {field}"}
        conn.commit(); conn.close()
    return {"ok":True}

class NewOrderRow(BaseModel):
    year_month: str
    order_seq: Optional[int] = None
    customer_name: Optional[str] = None
    item_desc: Optional[str] = None

@app.post("/api/orders/add-row")
def add_order_row(row: NewOrderRow):
    if is_supabase():
        if not row.order_seq:
            r = _sq("orders").select("order_seq").ilike("year_month",f"{row.year_month[:4]}%").order("order_seq",desc=True).limit(1).execute()
            row.order_seq = (r.data[0]["order_seq"]+1) if r.data else 1
        o = _sq("orders").insert({"year_month":row.year_month,"order_seq":row.order_seq,"customer_name":row.customer_name}).execute()
        order_id = o.data[0]["id"]
        item_id = None
        if row.item_desc:
            i = _sq("order_items").insert({"order_id":order_id,"item_desc":row.item_desc}).execute()
            item_id = i.data[0]["id"]
    else:
        conn = get_sqlite()
        if not row.order_seq:
            r = conn.execute("SELECT MAX(order_seq) FROM orders WHERE year_month LIKE ?",(row.year_month[:4]+"%",)).fetchone()
            row.order_seq = (r[0] or 0)+1
        cur = conn.execute("INSERT INTO orders (year_month,order_seq,customer_name) VALUES (?,?,?)",(row.year_month,row.order_seq,row.customer_name))
        order_id = cur.lastrowid
        item_id = None
        if row.item_desc:
            cur2 = conn.execute("INSERT INTO order_items (order_id,item_desc) VALUES (?,?)",(order_id,row.item_desc))
            item_id = cur2.lastrowid
        conn.commit(); conn.close()
    return {"ok":True,"order_id":order_id,"item_id":item_id,"order_seq":row.order_seq}

@app.post("/api/orders/add-item")
def add_order_item(order_id: int = Body(...), item_desc: str = Body("")):
    if is_supabase():
        r = _sq("order_items").insert({"order_id":order_id,"item_desc":item_desc}).execute()
        return {"ok":True,"item_id":r.data[0]["id"]}
    else:
        conn = get_sqlite()
        cur = conn.execute("INSERT INTO order_items (order_id,item_desc) VALUES (?,?)",(order_id,item_desc))
        item_id = cur.lastrowid; conn.commit(); conn.close()
        return {"ok":True,"item_id":item_id}

@app.post("/api/orders/delete-row")
def delete_order_row(order_id: int = Body(...), item_id: Optional[int] = Body(None)):
    if is_supabase():
        if item_id:
            _sq("order_items").delete().eq("id",item_id).execute()
            rem = _sq("order_items").select("id",count="exact").eq("order_id",order_id).execute()
            if rem.count == 0: _sq("orders").delete().eq("id",order_id).execute()
        else:
            _sq("order_items").delete().eq("order_id",order_id).execute()
            _sq("orders").delete().eq("id",order_id).execute()
    else:
        conn = get_sqlite()
        if item_id:
            conn.execute("DELETE FROM order_items WHERE id=?",(item_id,))
            rem = conn.execute("SELECT COUNT(*) FROM order_items WHERE order_id=?",(order_id,)).fetchone()[0]
            if rem == 0: conn.execute("DELETE FROM orders WHERE id=?",(order_id,))
        else:
            conn.execute("DELETE FROM order_items WHERE order_id=?",(order_id,))
            conn.execute("DELETE FROM orders WHERE id=?",(order_id,))
        conn.commit(); conn.close()
    return {"ok":True}


@app.post("/api/orders/delete-order")
def delete_entire_order(order_id: int = Body(...)):
    """Delete an entire order and all its items."""
    if is_supabase():
        _sq("order_items").delete().eq("order_id", order_id).execute()
        _sq("orders").delete().eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        conn.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
        conn.commit()
        conn.close()
    return {"ok": True}


@app.get("/api/orders/download")
def download_orders_excel(year_month: str = Query("")):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from urllib.parse import quote
    if is_supabase():
        oq = _sq("orders").select("*")
        if year_month: oq = oq.eq("year_month",year_month)
        orders = oq.order("order_seq").execute().data
        oids = [o["id"] for o in orders]
        all_items = []
        for i in range(0,len(oids),50):
            all_items.extend(_sq("order_items").select("*").in_("order_id",oids[i:i+50]).order("id").execute().data)
        items_map = {}
        for it in all_items: items_map.setdefault(it["order_id"],[]).append(it)
    else:
        conn = get_sqlite()
        orders = [dict(r) for r in conn.execute("SELECT * FROM orders WHERE year_month=? ORDER BY order_seq",(year_month,)).fetchall()] if year_month else [dict(r) for r in conn.execute("SELECT * FROM orders ORDER BY year_month DESC,order_seq").fetchall()]
        items_map = {}
        for o in orders: items_map[o["id"]] = [dict(r) for r in conn.execute("SELECT * FROM order_items WHERE order_id=? ORDER BY id",(o["id"],)).fetchall()]
        conn.close()

    wb = openpyxl.Workbook(); ws = wb.active
    title = f"{year_month.split('-')[0]}년도 {int(year_month.split('-')[1])}월 수주대장 (한국밸브 서울영업소)" if year_month else "수주대장"
    ws.title = year_month or "전체"
    thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    hf = PatternFill(start_color="D9E1F2",end_color="D9E1F2",fill_type="solid")
    ws.merge_cells("A1:T1"); ws["A1"]=title; ws["A1"].font=Font(bold=True,size=13); ws["A1"].alignment=Alignment(horizontal="center")
    hdrs = ["NO","업 체 명","특이사항","ITEM","단가(VAT제외)","수  량","수주일자","금액","재고판매","납기일자","납품일자","납품금액","매출일자","매출금액","REMARK","수금","매입업체","발주일자","매입금액","매입납기"]
    ws_w = [5,14,12,40,12,5,11,13,11,11,11,13,11,13,12,10,14,11,13,11]
    for i,(h,w) in enumerate(zip(hdrs,ws_w),1):
        c=ws.cell(row=3,column=i,value=h); c.font=Font(bold=True,size=9); c.fill=hf; c.border=thin; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    rn = 4
    for order in orders:
        items = items_map.get(order["id"],[])
        fi = items[0] if items else {}
        rd = [order.get("order_seq"),order.get("customer_name"),order.get("memo") or order.get("discount_rate"),fi.get("item_desc"),fi.get("unit_price"),fi.get("quantity"),order.get("order_date"),fi.get("amount") or order.get("total_amount"),order.get("stock_amount"),order.get("delivery_due"),order.get("delivery_date"),order.get("delivery_amount"),order.get("sales_date"),order.get("sales_amount"),order.get("remark"),order.get("collection_note"),order.get("supplier_name"),order.get("purchase_date"),order.get("purchase_amount"),order.get("purchase_due")]
        for i,v in enumerate(rd,1):
            c=ws.cell(row=rn,column=i,value=v); c.border=thin
            if isinstance(v,(int,float)) and v and i in(5,8,9,12,14,19): c.number_format='#,##0'
        rn+=1
        for it in items[1:]:
            ws.cell(row=rn,column=4,value=it.get("item_desc")).border=thin
            ws.cell(row=rn,column=5,value=it.get("unit_price")).border=thin
            ws.cell(row=rn,column=6,value=it.get("quantity")).border=thin
            if it.get("amount"): c=ws.cell(row=rn,column=8,value=it.get("amount")); c.border=thin; c.number_format='#,##0'
            rn+=1
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"수주대장_{year_month or '전체'}.xlsx"
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename*=UTF-8''{quote(fn)}"})

@app.post("/api/orders/upload")
async def upload_orders_excel(file: UploadFile = File(...), year_month: str = Query("")):
    import openpyxl
    if not year_month: return {"ok":False,"error":"year_month required"}
    content = await file.read(); wb = openpyxl.load_workbook(io.BytesIO(content),data_only=True); ws = wb.active
    header_row = None
    for r in range(1,min(10,ws.max_row+1)):
        for c in range(1,ws.max_column+1):
            v = ws.cell(r,c).value
            if v and ("ITEM" in str(v) or "업 체 명" in str(v)): header_row=r; break
        if header_row: break
    if not header_row: return {"ok":False,"error":"Cannot find header row"}
    col_map={}
    for c in range(1,ws.max_column+1):
        v=str(ws.cell(header_row,c).value or "").strip()
        if "NO"==v:col_map["seq"]=c
        elif "업 체 명" in v or "업체명" in v:col_map["customer"]=c
        elif "특이사항" in v:col_map["memo"]=c
        elif "ITEM" in v:col_map["item"]=c
        elif "단가" in v:col_map["unit_price"]=c
        elif "수  량" in v or "수량" in v:col_map["qty"]=c
        elif "수주일자" in v:col_map["order_date"]=c
        elif v=="금액" and "amount" not in col_map:col_map["amount"]=c
        elif "재고판매" in v:col_map["stock"]=c
        elif "납기일자" in v and "due" not in col_map:col_map["due"]=c
        elif "납품일자" in v:col_map["delivery_date"]=c
        elif "납품금액" in v:col_map["delivery_amount"]=c
        elif "매출일자" in v:col_map["sales_date"]=c
        elif "매출금액" in v:col_map["sales_amount"]=c
        elif "REMARK" in v.upper():col_map["remark"]=c
        elif "수금" in v:col_map["collection"]=c
        elif c>14 and "업체" in v:col_map["supplier"]=c
        elif c>14 and "발주일" in v:col_map["purchase_date"]=c
        elif c>14 and "금액" in v:col_map["purchase_amount"]=c
        elif c>14 and "납기" in v:col_map["purchase_due"]=c
    def cv(r,k): c=col_map.get(k); return ws.cell(r,c).value if c else None
    def tds(v):
        if v is None: return None
        if isinstance(v,datetime): return v.strftime("%Y-%m-%d")
        if isinstance(v,date): return v.isoformat()
        return str(v).strip() or None
    def ti(v):
        if v is None or v=="": return None
        try: return int(float(v))
        except (ValueError, TypeError): return None
    # Delete existing
    if is_supabase():
        ex=_sq("orders").select("id").eq("year_month",year_month).execute()
        if ex.data:
            eids=[r["id"] for r in ex.data]
            for i in range(0,len(eids),50): _sq("order_items").delete().in_("order_id",eids[i:i+50]).execute()
            _sq("orders").delete().eq("year_month",year_month).execute()
    else:
        conn=get_sqlite()
        oids=[r[0] for r in conn.execute("SELECT id FROM orders WHERE year_month=?",(year_month,)).fetchall()]
        if oids:
            ph=",".join("?"*len(oids)); conn.execute(f"DELETE FROM order_items WHERE order_id IN ({ph})",oids); conn.execute(f"DELETE FROM orders WHERE id IN ({ph})",oids)
        # No commit here — commit after all inserts so delete+insert is atomic
    imported=0; coid=None
    for r in range(header_row+1,ws.max_row+1):
        seq,cust,item=cv(r,"seq"),cv(r,"customer"),cv(r,"item")
        if not seq and not cust and not item: continue
        if cust and "TOTAL" in str(cust): continue
        if seq and isinstance(seq,(int,float)) and seq>0:
            mv=cv(r,"memo"); dr,ms=None,None
            if isinstance(mv,(int,float)) and 0<mv<1: dr=mv
            elif mv: ms=str(mv).strip() or None
            od={"year_month":year_month,"order_seq":int(seq),"customer_name":str(cust).strip() if cust else None,"memo":ms,"discount_rate":dr,"order_date":tds(cv(r,"order_date")),"total_amount":ti(cv(r,"amount")),"stock_amount":ti(cv(r,"stock")),"delivery_due":tds(cv(r,"due")),"delivery_date":tds(cv(r,"delivery_date")),"delivery_amount":ti(cv(r,"delivery_amount")),"sales_date":tds(cv(r,"sales_date")),"sales_amount":ti(cv(r,"sales_amount")),"remark":str(cv(r,"remark")).strip() if cv(r,"remark") else None,"collection_note":str(cv(r,"collection")).strip() if cv(r,"collection") else None,"supplier_name":str(cv(r,"supplier")).strip() if cv(r,"supplier") else None,"purchase_date":tds(cv(r,"purchase_date")),"purchase_amount":ti(cv(r,"purchase_amount")),"purchase_due":tds(cv(r,"purchase_due"))}
            if is_supabase(): res=_sq("orders").insert(od).execute(); coid=res.data[0]["id"]
            else: cur=conn.execute("INSERT INTO orders (year_month,order_seq,customer_name,memo,discount_rate,order_date,total_amount,stock_amount,delivery_due,delivery_date,delivery_amount,sales_date,sales_amount,remark,collection_note,supplier_name,purchase_date,purchase_amount,purchase_due) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",tuple(od.values())); coid=cur.lastrowid
            imported+=1
        if item and coid and str(item).strip():
            id_={"order_id":coid,"item_desc":str(item).strip(),"unit_price":ti(cv(r,"unit_price")),"quantity":ti(cv(r,"qty")),"amount":ti(cv(r,"amount"))}
            if is_supabase(): _sq("order_items").insert(id_).execute()
            else: conn.execute("INSERT INTO order_items (order_id,item_desc,unit_price,quantity,amount) VALUES (?,?,?,?,?)",tuple(id_.values()))
    if not is_supabase(): conn.commit(); conn.close()
    return {"ok":True,"imported":imported,"year_month":year_month}


# ============================================================
# API: Purchase Orders (발주목록)
# ============================================================

@app.get("/api/purchase-orders/years")
def po_years():
    if is_supabase():
        r=_sq("purchase_orders").select("year").execute()
        return sorted(set(x["year"] for x in r.data if x["year"]),reverse=True)
    else:
        conn=get_sqlite(); rows=conn.execute("SELECT DISTINCT year FROM purchase_orders ORDER BY year DESC").fetchall(); conn.close()
        return [r["year"] for r in rows]

@app.get("/api/purchase-orders/rows")
def list_po_rows(year: int = Query(0), q: str = Query("")):
    if is_supabase():
        qr=_sq("purchase_orders").select("*")
        if year: qr=qr.eq("year",year)
        if q: qr=qr.or_(f"po_number.ilike.%{q}%,item_desc.ilike.%{q}%,supplier_name.ilike.%{q}%")
        return qr.order("order_date",nullsfirst=False).order("id").execute().data
    else:
        conn=get_sqlite(); conds,params=[],[]
        if year: conds.append("year=?"); params.append(year)
        if q: conds.append("(po_number LIKE ? OR item_desc LIKE ? OR supplier_name LIKE ?)"); params.extend([f"%{q}%"]*3)
        w=" AND ".join(conds) if conds else "1=1"
        rows=[dict(r) for r in conn.execute(f"SELECT * FROM purchase_orders WHERE {w} ORDER BY order_date ASC NULLS LAST,id ASC",params).fetchall()]
        conn.close(); return rows

class POCellUpdate(BaseModel):
    po_id: int; field: str; value: Optional[str] = None

@app.post("/api/purchase-orders/update-cell")
def update_po_cell(update: POCellUpdate):
    valid={"po_number","supplier_name","item_desc","quantity","amount","order_date","delivery_due","delivery_date","shipped","customer_name","remark"}
    if update.field not in valid: return {"ok":False,"error":f"Unknown field: {update.field}"}
    value=update.value if update.value!="" else None
    if update.field in("quantity","amount") and value is not None:
        try: value=int(float(str(value).replace(",","")))
        except (ValueError, TypeError): pass
    if is_supabase(): _sq("purchase_orders").update({update.field:value}).eq("id",update.po_id).execute()
    else: col=_safe_col(update.field); conn=get_sqlite(); conn.execute(f"UPDATE purchase_orders SET {col}=? WHERE id=?",(value,update.po_id)); conn.commit(); conn.close()
    return {"ok":True}

class NewPORow(BaseModel):
    year: int; po_number: Optional[str]=None; supplier_name: Optional[str]=None; item_desc: Optional[str]=None

@app.post("/api/purchase-orders/add-row")
def add_po_row(row: NewPORow):
    d={"year":row.year,"po_number":row.po_number,"supplier_name":row.supplier_name,"item_desc":row.item_desc}
    if is_supabase(): r=_sq("purchase_orders").insert(d).execute(); return {"ok":True,"po_id":r.data[0]["id"]}
    else: conn=get_sqlite(); cur=conn.execute("INSERT INTO purchase_orders (year,po_number,supplier_name,item_desc) VALUES (?,?,?,?)",tuple(d.values())); conn.commit(); pid=cur.lastrowid; conn.close(); return {"ok":True,"po_id":pid}

@app.post("/api/purchase-orders/delete-row")
def delete_po_row(po_id: int = Body(...)):
    if is_supabase(): _sq("purchase_orders").delete().eq("id",po_id).execute()
    else: conn=get_sqlite(); conn.execute("DELETE FROM purchase_orders WHERE id=?",(po_id,)); conn.commit(); conn.close()
    return {"ok":True}

@app.get("/api/purchase-orders/download")
def download_po_excel(year: int = Query(0)):
    import openpyxl; from openpyxl.styles import Font,Alignment,Border,Side,PatternFill; from urllib.parse import quote
    if is_supabase():
        q=_sq("purchase_orders").select("*")
        if year: q=q.eq("year",year)
        rows=q.order("order_date",nullsfirst=False).order("id").execute().data
    else:
        conn=get_sqlite()
        rows=[dict(r) for r in conn.execute("SELECT * FROM purchase_orders WHERE year=? ORDER BY order_date ASC NULLS LAST,id ASC",(year,)).fetchall()] if year else [dict(r) for r in conn.execute("SELECT * FROM purchase_orders ORDER BY year DESC,order_date ASC NULLS LAST,id ASC").fetchall()]
        conn.close()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title=str(year) if year else "전체"
    thin=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    hf=PatternFill(start_color="D9E1F2",end_color="D9E1F2",fill_type="solid")
    ws.merge_cells("A1:I1"); ws["A1"]="매입발주목록"; ws["A1"].font=Font(bold=True,size=13); ws["A1"].alignment=Alignment(horizontal="center")
    hdrs=["주문번호","매입처","발주품목","Q'TY","금액(VAT제외)","발주일자","납기","출고여부","비고"]; ws_w=[20,14,50,6,14,12,14,10,16]
    for i,(h,w) in enumerate(zip(hdrs,ws_w),1):
        c=ws.cell(row=2,column=i,value=h); c.font=Font(bold=True,size=9); c.fill=hf; c.border=thin; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    for idx,row in enumerate(rows):
        r=idx+3; rd=[row.get("po_number"),row.get("supplier_name"),row.get("item_desc"),row.get("quantity"),row.get("amount"),row.get("order_date"),row.get("delivery_due"),row.get("shipped"),row.get("remark") or row.get("customer_name")]
        for i,v in enumerate(rd,1): c=ws.cell(row=r,column=i,value=v); c.border=thin; (setattr(c,'number_format','#,##0') if isinstance(v,(int,float)) and v and i in(4,5) else None)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); fn=f"발주목록_{year or '전체'}.xlsx"
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename*=UTF-8''{quote(fn)}"})

@app.post("/api/purchase-orders/upload")
async def upload_po_excel(file: UploadFile = File(...), year: int = Query(0)):
    import openpyxl
    if not year: return {"ok":False,"error":"year required"}
    content=await file.read(); wb=openpyxl.load_workbook(io.BytesIO(content),data_only=True); ws=wb.active
    hr=None
    for r in range(1,min(10,ws.max_row+1)):
        for c in range(1,ws.max_column+1):
            v=ws.cell(r,c).value
            if v and ("주문번호" in str(v) or "발주서" in str(v)): hr=r; break
        if hr: break
    if not hr: return {"ok":False,"error":"Cannot find header row"}
    cm={}
    for c in range(1,ws.max_column+1):
        v=str(ws.cell(hr,c).value or "").strip()
        if "주문번호" in v or "발주서" in v:cm["po"]=c
        elif "매입처" in v:cm["sup"]=c
        elif "발주품목" in v or "DESCRIPTION" in v:cm["item"]=c
        elif "QTY" in v.upper():cm["qty"]=c
        elif "금액" in v:cm["amt"]=c
        elif "발주일자" in v:cm["od"]=c
        elif "납기" in v:cm["dd"]=c
        elif "출고" in v:cm["sh"]=c
        elif "비고" in v:cm["rm"]=c
    def cv(r,k): c=cm.get(k); return ws.cell(r,c).value if c else None
    def tds(v):
        if v is None: return None
        if isinstance(v,datetime): return v.strftime("%Y-%m-%d")
        if isinstance(v,date): return v.isoformat()
        return str(v).strip() or None
    def ti(v):
        if v is None or v=="": return None
        try: return int(float(v))
        except (ValueError, TypeError): return None
    if is_supabase(): _sq("purchase_orders").delete().eq("year",year).execute()
    else: conn=get_sqlite(); conn.execute("DELETE FROM purchase_orders WHERE year=?",(year,))  # commit deferred until after inserts
    batch=[]; imported=0
    for r in range(hr+1,ws.max_row+1):
        po,item,amt=cv(r,"po"),cv(r,"item"),cv(r,"amt")
        if not po and not item and not amt: continue
        dd=cv(r,"dd"); dds=tds(dd) or (str(dd).strip() if dd else None)
        batch.append({"year":year,"po_number":str(po).strip() if po else None,"supplier_name":str(cv(r,"sup")).strip() if cv(r,"sup") else None,"item_desc":str(item).strip() if item else None,"quantity":ti(cv(r,"qty")),"amount":ti(amt),"order_date":tds(cv(r,"od")),"delivery_due":dds,"shipped":str(cv(r,"sh")).strip() if cv(r,"sh") else None,"remark":str(cv(r,"rm")).strip() if cv(r,"rm") else None})
        imported+=1
    if is_supabase() and batch:
        for i in range(0,len(batch),200): _sq("purchase_orders").insert(batch[i:i+200]).execute()
    elif batch:
        for rd in batch: conn.execute("INSERT INTO purchase_orders (year,po_number,supplier_name,item_desc,quantity,amount,order_date,delivery_due,shipped,remark) VALUES (?,?,?,?,?,?,?,?,?,?)",tuple(rd.values()))
        conn.commit(); conn.close()
    return {"ok":True,"imported":imported,"year":year}


# ============================================================
# API: Purchases (매입 내역)
# ============================================================

@app.get("/api/purchases")
def list_purchases(year_month: str = Query(""), q: str = Query(""), page: int = Query(1), per_page: int = Query(50)):
    off=(page-1)*per_page
    if is_supabase():
        qr=_sq("purchases").select("*",count="exact")
        if year_month: qr=qr.eq("year_month",year_month)
        if q: qr=qr.or_(f"po_ref.ilike.%{q}%,category.ilike.%{q}%,memo.ilike.%{q}%")
        r=qr.order("year_month",desc=True).order("id",desc=True).range(off,off+per_page-1).execute()
        return {"total":r.count,"page":page,"per_page":per_page,"items":r.data}
    else:
        conn=get_sqlite(); conds,params=[],[]
        if year_month: conds.append("year_month=?"); params.append(year_month)
        if q: conds.append("(po_ref LIKE ? OR category LIKE ? OR memo LIKE ?)"); params.extend([f"%{q}%"]*3)
        w=" AND ".join(conds) if conds else "1=1"
        total=conn.execute(f"SELECT COUNT(*) FROM purchases WHERE {w}",params).fetchone()[0]
        rows=[dict(r) for r in conn.execute(f"SELECT * FROM purchases WHERE {w} ORDER BY year_month DESC,id DESC LIMIT ? OFFSET ?",params+[per_page,off]).fetchall()]
        conn.close(); return {"total":total,"page":page,"per_page":per_page,"items":rows}

@app.get("/api/purchases/months")
def purchase_months():
    if is_supabase():
        r=_sq("purchases").select("year_month").execute()
        return sorted(set(x["year_month"] for x in r.data if x["year_month"]),reverse=True)
    else:
        conn=get_sqlite(); rows=conn.execute("SELECT DISTINCT year_month FROM purchases ORDER BY year_month DESC").fetchall(); conn.close()
        return [r["year_month"] for r in rows]

@app.get("/api/purchases/summary")
def purchase_summary(year_month: str = Query("")):
    if is_supabase():
        qr=_sq("purchases").select("purchase_amount,sales_amount,profit_half")
        if year_month: qr=qr.eq("year_month",year_month)
        data=qr.execute().data
        return {"cnt":len(data),"total_purchase":sum(r["purchase_amount"] or 0 for r in data),"total_sales":sum(r["sales_amount"] or 0 for r in data),"total_profit":sum(r["profit_half"] or 0 for r in data)}
    else:
        conn=get_sqlite(); w="WHERE year_month=?" if year_month else ""; p=(year_month,) if year_month else ()
        row=conn.execute(f"SELECT COUNT(*) as cnt,SUM(purchase_amount) as total_purchase,SUM(sales_amount) as total_sales,SUM(profit_half) as total_profit FROM purchases {w}",p).fetchone()
        conn.close(); return dict(row)

@app.get("/api/stats")
def dashboard_stats():
    if is_supabase():
        return {
            "orders":_sq("orders").select("id",count="exact").execute().count,
            "purchase_orders":_sq("purchase_orders").select("id",count="exact").execute().count,
            "purchases":_sq("purchases").select("id",count="exact").execute().count,
            "price_items":_sq("price_catalog").select("id",count="exact").execute().count,
        }
    else:
        conn=get_sqlite()
        r={"orders":conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0],"purchase_orders":conn.execute("SELECT COUNT(*) FROM purchase_orders").fetchone()[0],"purchases":conn.execute("SELECT COUNT(*) FROM purchases").fetchone()[0]}
        conn.close()
        conn2=get_sqlite("한국밸브_협가표.db"); r["price_items"]=conn2.execute("SELECT COUNT(*) FROM price_list_items").fetchone()[0]; conn2.close()
        return r

# ============================================================
# API: Unified Order Workflow (거래 상태 관리)
# ============================================================

VALID_TRANSITIONS = {
    "견적": ["수주확정", "취소"],
    "수주확정": ["발주중", "출고완료", "취소"],
    "발주중": ["입고완료", "취소"],
    "입고완료": ["출고완료", "반품", "교환"],
    "출고완료": ["정산완료", "반품", "교환"],
    "정산완료": [],
    "반품": [],
    "교환": [],
    "취소": [],
}


def _validate_transition(current: str, target: str):
    allowed = VALID_TRANSITIONS.get(current, [])
    if target not in allowed:
        raise HTTPException(400, f"'{current}' → '{target}' 전이 불가. 허용: {allowed}")


class CreateOrderRequest(BaseModel):
    spec_text: str
    customer_name: Optional[str] = None
    quantity: int = 1
    discount_rate: Optional[str] = None
    note: Optional[str] = None
    unit_price: Optional[int] = None


@app.post("/api/workflow/orders")
def create_order(req: CreateOrderRequest):
    """Create a new order from quote spec. Auto-looks up price from 협가표."""
    from datetime import datetime as dt

    # Auto-lookup price
    parsed = parse_query(req.spec_text)
    if req.discount_rate:
        parsed["discount_rate"] = req.discount_rate

    unit_price = None
    if is_supabase():
        query = _sq("price_catalog").select("unit_price")
        if parsed.get("product_type"):
            query = query.eq("product_type", parsed["product_type"])
        if parsed.get("pressure_class"):
            query = query.eq("pressure_class", parsed["pressure_class"])
        if parsed.get("size_a"):
            query = query.eq("size_a", parsed["size_a"])
        if parsed.get("discount_rate"):
            query = query.eq("discount_rate", parsed["discount_rate"])
        result = query.execute().data
        if len(result) == 1:
            unit_price = result[0]["unit_price"]
    else:
        conn = get_sqlite("한국밸브_협가표.db")
        conditions, params = [], []
        for key in ("product_type", "pressure_class", "size_a", "discount_rate"):
            if parsed.get(key):
                conditions.append(f"{key} = ?")
                params.append(parsed[key])
        if conditions:
            row = conn.execute(
                f"SELECT unit_price FROM price_list_items WHERE {' AND '.join(conditions)} LIMIT 1",
                params
            ).fetchone()
            if row:
                unit_price = row["unit_price"]
        conn.close()

    if req.unit_price is not None:
        unit_price = req.unit_price
    revenue = (unit_price * req.quantity) if unit_price else None

    order_data = {
        "state": "견적",
        "spec_text": req.spec_text,
        "customer_name": req.customer_name,
        "quantity": req.quantity,
        "discount_rate": float(req.discount_rate.replace("%", "")) / 100 if req.discount_rate and "%" in req.discount_rate else None,
        "unit_price": unit_price,
        "revenue": revenue,
        "note": req.note,
        "quoted_at": dt.now().isoformat(),
        "year_month": dt.now().strftime("%Y-%m"),
    }

    if is_supabase():
        # Auto-assign order_seq (max seq in this month + 1)
        ym = order_data["year_month"]
        max_seq = _sq("orders").select("order_seq").eq("year_month", ym).not_.is_("order_seq", "null").order("order_seq", desc=True).limit(1).execute().data
        next_seq = (max_seq[0]["order_seq"] + 1) if max_seq else 1
        order_data["order_seq"] = next_seq
        order_data["order_date"] = dt.now().strftime("%Y-%m-%d")

        r = _sq("orders").insert(order_data).execute()
        order = r.data[0]
        order_id = order["id"]

        # Also create order_item so it shows in 수주대장 spreadsheet
        _sq("order_items").insert({
            "order_id": order_id,
            "item_desc": req.spec_text,
            "unit_price": unit_price,
            "quantity": req.quantity,
            "amount": revenue,
        }).execute()
    else:
        conn = get_sqlite()
        _run_sqlite_migration(conn)
        cols = [k for k in order_data.keys()]
        vals = [order_data[k] for k in cols]
        placeholders = ",".join(["?"] * len(cols))
        cur = conn.execute(
            f"INSERT INTO orders ({','.join(cols)}) VALUES ({placeholders})", vals
        )
        order_data["id"] = cur.lastrowid
        conn.commit()
        conn.close()
        order = order_data

    _log(order.get("id") or order_data.get("id"), "견적 등록", f"{req.spec_text} | {req.customer_name} | {req.quantity}EA | {unit_price} KRW")

    return {
        "ok": True,
        "order": order,
        "parsed": {k: v for k, v in parsed.items() if v is not None and v != []},
        "price_found": unit_price is not None,
    }


@app.post("/api/workflow/orders/{order_id}/confirm")
def confirm_order(order_id: int):
    """수주확정: 견적 → 수주확정. Auto-checks inventory."""
    from datetime import datetime as dt

    if is_supabase():
        r = _sq("orders").select("state,spec_text").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        order = r.data[0]
        _validate_transition(order["state"], "수주확정")

        # Check inventory
        stock_available = _check_stock(order.get("spec_text", ""))

        update = {
            "state": "수주확정",
            "ordered_at": dt.now().isoformat(),
            "type": "재고품" if stock_available else "제조품",
        }
        _sq("orders").update(update).eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        row = conn.execute("SELECT state, spec_text FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(404, "Order not found")
        _validate_transition(row["state"], "수주확정")
        stock_available = _check_stock(row["spec_text"] or "")
        conn.execute(
            "UPDATE orders SET state=?, ordered_at=?, type=? WHERE id=?",
            ("수주확정", dt.now().isoformat(), "재고품" if stock_available else "제조품", order_id)
        )
        conn.commit()
        conn.close()

    _log(order_id, "수주 확정", f"유형: {'재고품' if stock_available else '제조품'} | 재고: {'있음' if stock_available else '없음'}")
    return {"ok": True, "state": "수주확정", "type": "재고품" if stock_available else "제조품", "stock_available": stock_available}


class IssuePORequest(BaseModel):
    manufacturer: str = "한국밸브"
    cost_unit_price: Optional[int] = None
    requested_delivery_at: Optional[str] = None


@app.post("/api/workflow/orders/{order_id}/issue-po")
def issue_po(order_id: int, req: IssuePORequest):
    """제조사 발주: 수주확정 → 발주중"""
    from datetime import datetime as dt

    if is_supabase():
        r = _sq("orders").select("state,quantity").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        _validate_transition(r.data[0]["state"], "발주중")
        cost = (req.cost_unit_price * r.data[0]["quantity"]) if req.cost_unit_price else None
        update = {
            "state": "발주중",
            "manufacturer": req.manufacturer,
            "po_issued_at": dt.now().isoformat(),
            "requested_delivery_at": req.requested_delivery_at,
            "cost": cost,
            "profit": (r.data[0].get("revenue") or 0) - (cost or 0) if cost else None,
        }
        _sq("orders").update(update).eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        row = conn.execute("SELECT state, quantity, revenue FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(404, "Order not found")
        _validate_transition(row["state"], "발주중")
        cost = (req.cost_unit_price * row["quantity"]) if req.cost_unit_price else None
        profit = (row["revenue"] or 0) - (cost or 0) if cost else None
        conn.execute(
            "UPDATE orders SET state=?, manufacturer=?, po_issued_at=?, requested_delivery_at=?, cost=?, profit=? WHERE id=?",
            ("발주중", req.manufacturer, dt.now().isoformat(), req.requested_delivery_at, cost, profit, order_id)
        )
        conn.commit()
        conn.close()

    _log(order_id, "제조사 발주", f"발주처: {req.manufacturer} | 매입단가: {req.cost_unit_price} | 납기: {req.requested_delivery_at}")
    return {"ok": True, "state": "발주중"}


class ReceiveRequest(BaseModel):
    manufacturer_invoice_number: Optional[str] = None
    manufacturer_invoice_amount: Optional[int] = None


@app.post("/api/workflow/orders/{order_id}/receive")
def receive_order(order_id: int, req: ReceiveRequest):
    """입고 체크: 발주중 → 입고완료. Auto-matches invoice amount vs PO cost."""
    from datetime import datetime as dt

    match_status = "미입력"

    if is_supabase():
        r = _sq("orders").select("state,cost,manufacturer").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        order = r.data[0]
        _validate_transition(order["state"], "입고완료")

        if req.manufacturer_invoice_amount is not None and order.get("cost"):
            match_status = "일치" if req.manufacturer_invoice_amount == order["cost"] else "불일치"

        update = {
            "state": "입고완료",
            "received_at": dt.now().isoformat(),
            "manufacturer_invoice_number": req.manufacturer_invoice_number,
            "manufacturer_invoice_amount": req.manufacturer_invoice_amount,
        }
        _sq("orders").update(update).eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        row = conn.execute("SELECT state, cost FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(404, "Order not found")
        _validate_transition(row["state"], "입고완료")

        if req.manufacturer_invoice_amount is not None and row["cost"]:
            match_status = "일치" if req.manufacturer_invoice_amount == row["cost"] else "불일치"

        conn.execute(
            "UPDATE orders SET state=?, received_at=?, manufacturer_invoice_number=?, manufacturer_invoice_amount=? WHERE id=?",
            ("입고완료", dt.now().isoformat(), req.manufacturer_invoice_number, req.manufacturer_invoice_amount, order_id)
        )
        conn.commit()
        conn.close()

    diff = None
    if req.manufacturer_invoice_amount is not None and match_status == "불일치":
        cost = (r.data[0]["cost"] if is_supabase() else row["cost"]) or 0
        diff = req.manufacturer_invoice_amount - cost

    _log(order_id, "입고 완료", f"거래명세: {req.manufacturer_invoice_number} | 금액: {req.manufacturer_invoice_amount} | 매칭: {match_status}")
    return {"ok": True, "state": "입고완료", "match_status": match_status, "difference": diff}


@app.post("/api/workflow/orders/{order_id}/ship")
def ship_order(order_id: int):
    """고객 출고: 수주확정(재고) or 입고완료 → 출고완료"""
    from datetime import datetime as dt

    if is_supabase():
        r = _sq("orders").select("state,type,revenue,cost,quantity,unit_price").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        order = r.data[0]
        _validate_transition(order["state"], "출고완료")

        update = {"state": "출고완료", "shipped_at": dt.now().isoformat()}
        # For 재고품, compute cost from stock cost (simplified: same as unit_price * 0.5 placeholder)
        if order["type"] == "재고품" and not order.get("cost"):
            update["cost"] = order.get("revenue")  # placeholder — actual cost from ERP
        _sq("orders").update(update).eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        row = conn.execute("SELECT state, type, revenue FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(404, "Order not found")
        _validate_transition(row["state"], "출고완료")
        conn.execute(
            "UPDATE orders SET state=?, shipped_at=? WHERE id=?",
            ("출고완료", dt.now().isoformat(), order_id)
        )
        conn.commit()
        conn.close()

    # Deduct inventory if 재고품
    if is_supabase():
        full = _sq("orders").select("type,spec_text,quantity").eq("id", order_id).execute().data
        if full and full[0].get("type") == "재고품" and full[0].get("spec_text"):
            from parser import parse_query
            parsed = parse_query(full[0]["spec_text"])
            pt, pc, sv = parsed.get("product_type"), parsed.get("pressure_class"), parsed.get("size_a")
            if pt:
                inv_q = _sq("erp_inventory").select("id,stock_quantity").eq("product_type", pt)
                if pc: inv_q = inv_q.eq("pressure_class", pc)
                if sv:
                    sv_num = sv.replace("A", "")
                    inv_q = inv_q.eq("size_value", sv_num)
                inv = inv_q.limit(1).execute().data
                if inv and inv[0]["stock_quantity"] > 0:
                    new_qty = max(0, inv[0]["stock_quantity"] - (full[0].get("quantity") or 1))
                    _sq("erp_inventory").update({"stock_quantity": new_qty}).eq("id", inv[0]["id"]).execute()
                    _log(order_id, "재고 차감", f"{pt} {pc} {sv}: {inv[0]['stock_quantity']} → {new_qty}")

    _log(order_id, "출고 완료", None)
    return {"ok": True, "state": "출고완료"}


class CancelRequest(BaseModel):
    target_state: str = "취소"  # 취소, 반품, 교환


@app.post("/api/workflow/orders/{order_id}/cancel")
def cancel_order(order_id: int, req: CancelRequest):
    """취소/반품/교환 처리"""
    if req.target_state not in ("취소", "반품", "교환"):
        raise HTTPException(400, f"Invalid target: {req.target_state}")

    if is_supabase():
        r = _sq("orders").select("state").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        _validate_transition(r.data[0]["state"], req.target_state)
        _sq("orders").update({"state": req.target_state}).eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        row = conn.execute("SELECT state FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(404, "Order not found")
        _validate_transition(row["state"], req.target_state)
        conn.execute("UPDATE orders SET state=? WHERE id=?", (req.target_state, order_id))
        conn.commit()
        conn.close()

    _log(order_id, req.target_state, None)
    return {"ok": True, "state": req.target_state}


# ============================================================
# API: Workflow Views (자동 생성 뷰)
# ============================================================

@app.get("/api/workflow/views/수주대장")
def view_order_book(year_month: str = Query("")):
    """수주대장 view: confirmed+ orders grouped by month."""
    if is_supabase():
        q = _sq("orders").select("*").neq("state", "견적")
        if year_month:
            q = q.eq("year_month", year_month)
        return q.order("ordered_at", desc=True).execute().data
    else:
        conn = get_sqlite()
        w = "state != '견적'"
        params = []
        if year_month:
            w += " AND year_month = ?"
            params.append(year_month)
        try:
            rows = [dict(r) for r in conn.execute(
                f"SELECT * FROM orders WHERE {w} ORDER BY ordered_at DESC", params
            ).fetchall()]
        except Exception:
            rows = []
        conn.close()
        return rows


@app.get("/api/workflow/views/발주목록")
def view_po_list(year_month: str = Query("")):
    """발주목록 view: orders in 발주중 or 입고완료 state."""
    if is_supabase():
        q = _sq("orders").select("*").in_("state", ["발주중", "입고완료"])
        if year_month:
            q = q.eq("year_month", year_month)
        return q.order("po_issued_at", desc=True).execute().data
    else:
        conn = get_sqlite()
        w = "state IN ('발주중','입고완료')"
        params = []
        if year_month:
            w += " AND year_month = ?"
            params.append(year_month)
        try:
            rows = [dict(r) for r in conn.execute(
                f"SELECT * FROM orders WHERE {w} ORDER BY po_issued_at DESC", params
            ).fetchall()]
        except Exception:
            rows = []
        conn.close()
        return rows


@app.get("/api/workflow/views/매입내역")
def view_purchases(year_month: str = Query("")):
    """매입내역 view: orders with revenue/cost/profit."""
    if is_supabase():
        q = _sq("orders").select("*").in_("state", ["입고완료", "출고완료", "정산완료"])
        if year_month:
            q = q.eq("year_month", year_month)
        return q.order("received_at", desc=True).execute().data
    else:
        conn = get_sqlite()
        w = "state IN ('입고완료','출고완료','정산완료')"
        params = []
        if year_month:
            w += " AND year_month = ?"
            params.append(year_month)
        try:
            rows = [dict(r) for r in conn.execute(
                f"SELECT * FROM orders WHERE {w} ORDER BY received_at DESC", params
            ).fetchall()]
        except Exception:
            rows = []
        conn.close()
        return rows


@app.get("/api/workflow/orders")
def list_workflow_orders(state: str = Query(""), q: str = Query("")):
    """List all workflow orders with optional state/text filter."""
    if is_supabase():
        query = _sq("orders").select("*").not_.is_("state", "null")
        if state:
            query = query.eq("state", state)
        if q:
            query = query.or_(f"spec_text.ilike.%{q}%,customer_name.ilike.%{q}%")
        return query.order("created_at", desc=True).limit(1000).execute().data
    else:
        conn = get_sqlite()
        conds = ["state IS NOT NULL"]
        params = []
        if state:
            conds.append("state = ?")
            params.append(state)
        if q:
            conds.append("(spec_text LIKE ? OR customer_name LIKE ?)")
            params.extend([f"%{q}%", f"%{q}%"])
        w = " AND ".join(conds)
        try:
            rows = [dict(r) for r in conn.execute(
                f"SELECT * FROM orders WHERE {w} ORDER BY created_at DESC LIMIT 200", params
            ).fetchall()]
        except Exception:
            rows = []
        conn.close()
        return rows


# ============================================================
# API: ERP Inventory Import (재고 동기화)
# ============================================================

@app.post("/api/workflow/inventory/import")
async def import_erp_inventory(file: UploadFile = File(...)):
    """Import ERP inventory Excel (재고변동표 format)."""
    import openpyxl
    from datetime import datetime as dt

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Find header row (품목코드)
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value and "품목코드" in str(ws.cell(r, c).value):
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        return {"ok": False, "error": "Cannot find header row (품목코드)"}

    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or code == "합계":
            break
        product_type = ws.cell(r, 2).value
        pressure_class = str(ws.cell(r, 3).value) if ws.cell(r, 3).value else None
        stock_qty = ws.cell(r, 7).value  # 재고수량

        # Parse size from code (e.g., GA-10K-100 → 100)
        parts = str(code).split("-")
        size_value = parts[-1] if len(parts) >= 3 else None

        # Normalize product_type
        type_map = {"Y-STR": "Y-STRAINER"}
        normalized_type = type_map.get(product_type, product_type)

        items.append({
            "item_code": str(code).strip(),
            "product_type": normalized_type,
            "pressure_class": pressure_class,
            "size_value": size_value,
            "stock_quantity": int(stock_qty) if stock_qty else 0,
        })

    if not items:
        return {"ok": False, "error": "No inventory data found"}

    if is_supabase():
        # Upsert: delete all then insert
        _sq("erp_inventory").delete().neq("id", 0).execute()
        for i in range(0, len(items), 200):
            _sq("erp_inventory").insert(items[i:i+200]).execute()
    else:
        conn = get_sqlite()
        conn.execute("""CREATE TABLE IF NOT EXISTS erp_inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_code TEXT NOT NULL UNIQUE,
            product_type TEXT,
            pressure_class TEXT,
            size_value TEXT,
            stock_quantity INTEGER DEFAULT 0,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        conn.execute("DELETE FROM erp_inventory")
        for item in items:
            conn.execute(
                "INSERT INTO erp_inventory (item_code, product_type, pressure_class, size_value, stock_quantity) VALUES (?,?,?,?,?)",
                (item["item_code"], item["product_type"], item["pressure_class"], item["size_value"], item["stock_quantity"])
            )
        conn.commit()
        conn.close()

    return {"ok": True, "imported": len(items), "total_stock": sum(i["stock_quantity"] for i in items)}


@app.get("/api/workflow/inventory")
def get_inventory(q: str = Query("")):
    """Get current ERP inventory."""
    if is_supabase():
        query = _sq("erp_inventory").select("*")
        if q:
            query = query.or_(f"item_code.ilike.%{q}%,product_type.ilike.%{q}%")
        return query.order("product_type").order("pressure_class").order("size_value").execute().data
    else:
        conn = get_sqlite()
        try:
            if q:
                rows = [dict(r) for r in conn.execute(
                    "SELECT * FROM erp_inventory WHERE item_code LIKE ? OR product_type LIKE ? ORDER BY product_type, pressure_class, size_value",
                    (f"%{q}%", f"%{q}%")
                ).fetchall()]
            else:
                rows = [dict(r) for r in conn.execute(
                    "SELECT * FROM erp_inventory ORDER BY product_type, pressure_class, size_value"
                ).fetchall()]
        except Exception:
            rows = []
        conn.close()
        return rows


def _check_stock(spec_text: str) -> bool:
    """Check if item is available in ERP inventory based on spec text."""
    parsed = parse_query(spec_text)
    product_type = parsed.get("product_type")
    pressure_class = parsed.get("pressure_class")
    size_a = parsed.get("size_a")

    if not product_type or not size_a:
        return False

    # Convert size_a (e.g., "80A") to numeric
    size_num = size_a.replace("A", "").replace("a", "")

    if is_supabase():
        q = _sq("erp_inventory").select("stock_quantity").eq("product_type", product_type)
        if pressure_class:
            q = q.eq("pressure_class", pressure_class)
        q = q.eq("size_value", size_num)
        r = q.execute().data
        return bool(r and r[0]["stock_quantity"] > 0)
    else:
        conn = get_sqlite()
        try:
            conds = ["product_type = ?", "size_value = ?"]
            params = [product_type, size_num]
            if pressure_class:
                conds.append("pressure_class = ?")
                params.append(pressure_class)
            row = conn.execute(
                f"SELECT stock_quantity FROM erp_inventory WHERE {' AND '.join(conds)} LIMIT 1",
                params
            ).fetchone()
            conn.close()
            return bool(row and row["stock_quantity"] > 0)
        except:
            conn.close()
            return False


# ============================================================
# API: PDF Generation (견적서, 발주서, 거래명세표)
# ============================================================

from web.pdf_generator import generate_quotation_pdf, generate_po_pdf, generate_invoice_pdf


def _get_workflow_order(order_id: int) -> dict:
    """Fetch a single workflow order by ID."""
    if is_supabase():
        r = _sq("orders").select("*").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        return r.data[0]
    else:
        conn = get_sqlite()
        _run_sqlite_migration(conn)
        row = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        conn.close()
        if not row:
            raise HTTPException(404, "Order not found")
        return dict(row)


@app.get("/api/workflow/orders/{order_id}/pdf/{doc_type}")
def download_pdf(order_id: int, doc_type: str):
    """Download PDF for an order. doc_type: 견적서, 발주서, 거래명세표"""
    from urllib.parse import quote
    order = _get_workflow_order(order_id)

    if doc_type == "견적서":
        buf = generate_quotation_pdf(order)
        fn = f"견적서_MSM-Q-{order_id}.pdf"
    elif doc_type == "발주서":
        buf = generate_po_pdf(order)
        fn = f"발주서_MSM-PO-{order_id}.pdf"
    elif doc_type == "거래명세표":
        buf = generate_invoice_pdf([order])
        fn = f"거래명세표_{order_id}.pdf"
    else:
        raise HTTPException(400, f"Unknown doc type: {doc_type}")

    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fn)}"}
    )


@app.get("/api/workflow/orders/{order_id}/xlsx/견적서")
def download_quotation_xlsx(order_id: int):
    """Download editable .xlsx quotation using the 견적문의서 template."""
    import openpyxl, copy, shutil
    from urllib.parse import quote as url_quote
    from datetime import datetime as dt

    order = _get_workflow_order(order_id)

    # Copy template
    template_path = os.path.join(BASE_DIR, "templates", "견적문의서.xlsx")
    if not os.path.exists(template_path):
        # Fallback: try Desktop
        template_path = os.path.expanduser("~/Desktop/견적문의서.xlsx")
    if not os.path.exists(template_path):
        raise HTTPException(404, "견적문의서 템플릿 파일을 찾을 수 없습니다")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Fill in data
    # A3: date
    ws["A3"] = dt.now().strftime("%Y-%m-%d")

    # A7: 견적번호
    quote_num = f"MSM-Q-{order_id}"
    ws["A7"] = f" 견적번호 : {quote_num}"

    # B4: customer (channel/distributor)
    ws["B4"] = order.get("customer_name") or ""

    # Item rows start at row 16
    spec = order.get("spec_text") or ""
    qty = order.get("quantity") or 1
    price = order.get("unit_price") or 0

    # Row 16: first item
    ws.cell(row=16, column=3, value=spec)    # C16: Description
    ws.cell(row=16, column=7, value=qty)     # G16: Qty
    ws.cell(row=16, column=8, value=price)   # H16: Unit Price
    # I16 already has formula =G16*H16

    # C18: MAKER (keep default 한국밸브 or set from order)
    manufacturer = order.get("manufacturer") or "한국밸브"
    ws["C18"] = f"MAKER : {manufacturer}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"견적서_{quote_num}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{url_quote(fn)}"},
    )


class BatchInvoiceRequest(BaseModel):
    order_ids: list[int]


@app.post("/api/workflow/pdf/거래명세표")
def download_batch_invoice(req: BatchInvoiceRequest):
    """Download batch transaction statement PDF for multiple orders."""
    from urllib.parse import quote
    orders = []
    for oid in req.order_ids:
        orders.append(_get_workflow_order(oid))
    if not orders:
        raise HTTPException(400, "No orders")

    buf = generate_invoice_pdf(orders)
    fn = f"거래명세표_일괄_{len(orders)}건.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fn)}"}
    )


# ============================================================
# API: 거래명세 검증 리스트
# ============================================================

@app.get("/api/workflow/views/거래명세검증")
def view_invoice_verification(year_month: str = Query("")):
    """거래명세 검증: orders with manufacturer_invoice_number, showing match status."""
    if is_supabase():
        q = _sq("orders").select("id,customer_name,spec_text,cost,manufacturer_invoice_number,manufacturer_invoice_amount,manufacturer,state")
        q = q.not_.is_("manufacturer_invoice_number", "null")
        if year_month:
            q = q.eq("year_month", year_month)
        rows = q.order("received_at", desc=True).execute().data
    else:
        conn = get_sqlite()
        _run_sqlite_migration(conn)
        w = "manufacturer_invoice_number IS NOT NULL"
        params = []
        if year_month:
            w += " AND year_month = ?"
            params.append(year_month)
        try:
            rows = [dict(r) for r in conn.execute(
                f"SELECT id, customer_name, spec_text, cost, manufacturer_invoice_number, manufacturer_invoice_amount, manufacturer, state FROM orders WHERE {w} ORDER BY received_at DESC",
                params
            ).fetchall()]
        except Exception:
            rows = []
        conn.close()

    # Add match_status to each row
    for r in rows:
        inv_amt = r.get("manufacturer_invoice_amount")
        cost = r.get("cost")
        if inv_amt is not None and cost is not None:
            r["match_status"] = "일치" if inv_amt == cost else "불일치"
            r["difference"] = inv_amt - cost
        else:
            r["match_status"] = "미입력"
            r["difference"] = None

    return rows


# ============================================================
# API: Workflow View Excel Download
# ============================================================

VIEW_COLUMNS = {
    "수주대장": [
        ("상태", "state"), ("고객", "customer_name"), ("사양", "spec_text"),
        ("수량", "quantity"), ("단가", "unit_price"), ("매출", "revenue"),
        ("수주일", "ordered_at"), ("출고일", "shipped_at"),
    ],
    "발주목록": [
        ("상태", "state"), ("고객", "customer_name"), ("사양", "spec_text"),
        ("발주처", "manufacturer"), ("수량", "quantity"), ("매입", "cost"),
        ("발주일", "po_issued_at"), ("납기", "requested_delivery_at"),
    ],
    "매입내역": [
        ("상태", "state"), ("고객", "customer_name"), ("사양", "spec_text"),
        ("매출", "revenue"), ("매입", "cost"), ("이익", "profit"),
        ("입고일", "received_at"),
    ],
    "거래명세검증": [
        ("고객", "customer_name"), ("사양", "spec_text"),
        ("발주금액", "cost"), ("거래명세번호", "manufacturer_invoice_number"),
        ("거래명세금액", "manufacturer_invoice_amount"), ("매칭", "match_status"),
    ],
}


@app.get("/api/workflow/views/{view_name}/download")
def download_view_excel(view_name: str, year_month: str = Query("")):
    """Download a workflow view as Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from urllib.parse import quote as url_quote

    cols = VIEW_COLUMNS.get(view_name)
    if not cols:
        raise HTTPException(400, f"Unknown view: {view_name}")

    # Reuse existing view endpoints to get data
    if view_name == "수주대장":
        rows = view_order_book(year_month=year_month)
    elif view_name == "발주목록":
        rows = view_po_list(year_month=year_month)
    elif view_name == "매입내역":
        rows = view_purchases(year_month=year_month)
    elif view_name == "거래명세검증":
        rows = view_invoice_verification(year_month=year_month)
    else:
        rows = []

    wb = openpyxl.Workbook()
    ws = wb.active
    month_label = f"{year_month}" if year_month else "전체"
    ws.title = month_label
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    hf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    title = f"{view_name} — {month_label}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    for i, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=3, column=i, value=label)
        c.font = Font(bold=True, size=9)
        c.fill = hf
        c.border = thin
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data
    num_fields = {"unit_price", "revenue", "cost", "profit", "quantity",
                  "manufacturer_invoice_amount"}
    for ri, row in enumerate(rows, 4):
        for ci, (_, key) in enumerate(cols, 1):
            val = row.get(key)
            if isinstance(val, str) and len(val) > 10 and "T" in val:
                val = val[:10]  # trim datetime to date
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            if key in num_fields and isinstance(val, (int, float)):
                c.number_format = "#,##0"

    # Auto-width
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"{view_name}_{month_label}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{url_quote(fn)}"},
    )


# ============================================================
# API: 월말 정산
# ============================================================

class SettlementRequest(BaseModel):
    year_month: str


@app.post("/api/workflow/settlement")
def run_settlement(req: SettlementRequest):
    """월말 정산: mark all 출고완료 orders in the month as 정산완료."""
    from datetime import datetime as dt

    if is_supabase():
        orders = _sq("orders").select("id,revenue,cost,profit,customer_name,spec_text,manufacturer_invoice_number,manufacturer_invoice_amount") \
            .eq("state", "출고완료").eq("year_month", req.year_month).execute().data
        if orders:
            ids = [o["id"] for o in orders]
            for i in range(0, len(ids), 50):
                _sq("orders").update({"state": "정산완료"}).in_("id", ids[i:i+50]).execute()
    else:
        conn = get_sqlite()
        _run_sqlite_migration(conn)
        try:
            orders = [dict(r) for r in conn.execute(
                "SELECT id, revenue, cost, profit, customer_name, spec_text, manufacturer_invoice_number, manufacturer_invoice_amount FROM orders WHERE state='출고완료' AND year_month=?",
                (req.year_month,)
            ).fetchall()]
            if orders:
                ids = [o["id"] for o in orders]
                conn.execute(f"UPDATE orders SET state='정산완료' WHERE id IN ({','.join('?' * len(ids))})", ids)
                conn.commit()
        except Exception:
            orders = []
        conn.close()

    total_revenue = sum(o.get("revenue") or 0 for o in orders)
    total_cost = sum(o.get("cost") or 0 for o in orders)
    total_profit = sum(o.get("profit") or 0 for o in orders)

    return {
        "ok": True,
        "year_month": req.year_month,
        "settled_count": len(orders),
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "total_profit": total_profit,
        "orders": orders,
    }


@app.get("/api/workflow/settlement/download")
def download_settlement_excel(year_month: str = Query("")):
    """Download 거래명세 매칭표 Excel for a month."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from urllib.parse import quote

    if not year_month:
        raise HTTPException(400, "year_month required")

    if is_supabase():
        orders = _sq("orders").select("*").eq("year_month", year_month) \
            .in_("state", ["출고완료", "정산완료"]).order("id").execute().data
    else:
        conn = get_sqlite()
        _run_sqlite_migration(conn)
        try:
            orders = [dict(r) for r in conn.execute(
                "SELECT * FROM orders WHERE year_month=? AND state IN ('출고완료','정산완료') ORDER BY id",
                (year_month,)
            ).fetchall()]
        except Exception:
            orders = []
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year_month} 매칭표"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    hf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells("A1:I1")
    ws["A1"] = f"{year_month} 거래명세 매칭표"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    hdrs = ["No", "고객사", "사양", "매출액", "매입액", "이익", "거래명세번호", "거래명세금액", "매칭"]
    ws_w = [6, 16, 40, 14, 14, 14, 16, 14, 8]
    for i, (h, w) in enumerate(zip(hdrs, ws_w), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, size=9)
        c.fill = hf
        c.border = thin
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for idx, o in enumerate(orders):
        r = idx + 4
        inv_amt = o.get("manufacturer_invoice_amount")
        cost = o.get("cost")
        if inv_amt is not None and cost is not None:
            match = "O" if inv_amt == cost else "X"
        else:
            match = "-"

        data = [idx + 1, o.get("customer_name"), o.get("spec_text"),
                o.get("revenue"), o.get("cost"), o.get("profit"),
                o.get("manufacturer_invoice_number"), inv_amt, match]
        for i, v in enumerate(data, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = thin
            if isinstance(v, (int, float)) and v and i in (4, 5, 6, 8):
                c.number_format = '#,##0'
            if i == 9:
                if match == "O":
                    c.fill = match_fill
                elif match == "X":
                    c.fill = mismatch_fill

    # Totals
    total_row = len(orders) + 4
    ws.cell(row=total_row, column=2, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=2).border = thin
    for col, key in [(4, "revenue"), (5, "cost"), (6, "profit")]:
        total = sum(o.get(key) or 0 for o in orders)
        c = ws.cell(row=total_row, column=col, value=total)
        c.font = Font(bold=True)
        c.number_format = '#,##0'
        c.border = thin

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"매칭표_{year_month}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fn)}"}
    )


# ============================================================
# API: Delete workflow order
# ============================================================

@app.post("/api/workflow/orders/{order_id}/delete")
def delete_workflow_order(order_id: int):
    """Delete a workflow order (only 견적 and 취소 states)."""
    if is_supabase():
        r = _sq("orders").select("state").eq("id", order_id).execute()
        if not r.data:
            raise HTTPException(404, "Order not found")
        if r.data[0]["state"] not in ("견적", "취소"):
            raise HTTPException(400, "확정된 거래는 삭제할 수 없습니다. 취소 처리 후 삭제하세요.")
        _sq("order_items").delete().eq("order_id", order_id).execute()
        _sq("orders").delete().eq("id", order_id).execute()
    else:
        conn = get_sqlite()
        row = conn.execute("SELECT state FROM orders WHERE id=?", (order_id,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(404, "Order not found")
        if row["state"] not in ("견적", "취소"):
            conn.close()
            raise HTTPException(400, "확정된 거래는 삭제할 수 없습니다.")
        conn.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
        conn.commit()
        conn.close()
    return {"ok": True}


_migration_done = False

def _run_sqlite_migration(conn):
    """Apply migration columns to SQLite (local dev only). Idempotent."""
    global _migration_done
    if _migration_done:
        return
    migrations = [
        "ALTER TABLE orders ADD COLUMN type TEXT",
        "ALTER TABLE orders ADD COLUMN state TEXT DEFAULT '견적'",
        "ALTER TABLE orders ADD COLUMN spec_text TEXT",
        "ALTER TABLE orders ADD COLUMN quantity INTEGER DEFAULT 1",
        "ALTER TABLE orders ADD COLUMN revenue INTEGER",
        "ALTER TABLE orders ADD COLUMN cost INTEGER",
        "ALTER TABLE orders ADD COLUMN profit INTEGER",
        "ALTER TABLE orders ADD COLUMN quoted_at TEXT",
        "ALTER TABLE orders ADD COLUMN ordered_at TEXT",
        "ALTER TABLE orders ADD COLUMN po_issued_at TEXT",
        "ALTER TABLE orders ADD COLUMN requested_delivery_at TEXT",
        "ALTER TABLE orders ADD COLUMN received_at TEXT",
        "ALTER TABLE orders ADD COLUMN shipped_at TEXT",
        "ALTER TABLE orders ADD COLUMN manufacturer TEXT",
        "ALTER TABLE orders ADD COLUMN manufacturer_invoice_number TEXT",
        "ALTER TABLE orders ADD COLUMN manufacturer_invoice_amount INTEGER",
        "ALTER TABLE orders ADD COLUMN note TEXT",
    ]
    for sql in migrations:
        try:
            conn.execute(sql)
        except Exception:
            pass  # Column already exists
    conn.commit()
    _migration_done = True


# ============================================================
# Static / Entry point
# ============================================================

@app.get("/", response_class=HTMLResponse)
def index():
    with open(os.path.join(STATIC_DIR,"index.html"),encoding="utf-8") as f: return f.read()

if __name__ == "__main__":
    import uvicorn; uvicorn.run("web.server:app",host="0.0.0.0",port=8000,reload=True)
